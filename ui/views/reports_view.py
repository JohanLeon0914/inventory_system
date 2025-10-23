"""
Vista de Reportes - Estadísticas y reportes del negocio
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QComboBox, QDateEdit,
    QHeaderView, QGroupBox, QFormLayout, QScrollArea, QFrame,
    QMessageBox, QFileDialog
)
from PyQt6.QtCore import Qt, QDate
from datetime import datetime, timedelta
from config.database import get_session, close_session
from models import Sale, Product, Customer, SaleItem, RawMaterial, RawMaterialMovement, RawMaterialMovementType, ProductMaterial
from sqlalchemy import func, desc
from sqlalchemy.orm import joinedload

class ReportsView(QWidget):
    """Vista para reportes y estadísticas"""
    def __init__(self):
        super().__init__()
        self.init_ui()
        # Cargar reportes después de que la UI esté lista
        self.load_reports()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Aplicar estilo general al widget
        self.setStyleSheet("""
            QWidget {
                background-color: #f8fafc;
            }
            QLabel {
                color: #0f172a;
            }
        """)
        
        # Header
        title = QLabel("Reportes y Estadísticas")
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #0f172a; background-color: transparent;")
        layout.addWidget(title)
        
        # Scroll area para el contenido
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        
        content_widget = QWidget()
        content_layout = QVBoxLayout(content_widget)
        content_layout.setSpacing(20)
        
        # Filtros de fecha
        filters_group = QGroupBox("Período de Reporte")
        filters_group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 15px;
                padding: 20px 15px 15px 15px;
                font-weight: bold;
                color: #0f172a;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 10px;
                top: 5px;
                padding: 0 5px;
                background-color: white;
                color: #0f172a;
            }
            QLabel {
                color: #0f172a;
                background-color: transparent;
            }
        """)
        filters_layout = QHBoxLayout()
        filters_layout.setContentsMargins(10, 25, 10, 10)
        
        filters_layout.addWidget(QLabel("Desde:"))
        self.date_from = QDateEdit()
        self.date_from.setDate(QDate.currentDate().addDays(-30))
        self.date_from.setCalendarPopup(True)
        self.date_from.setDisplayFormat("dd/MM/yyyy")
        self.date_from.setMinimumHeight(35)
        self.date_from.setStyleSheet("""
            QDateEdit {
                color: #0f172a;
                background-color: white;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                padding: 5px 10px;
            }
            QCalendarWidget QWidget {
                color: #0f172a;
            }
            QCalendarWidget QAbstractItemView {
                color: #0f172a;
                background-color: white;
                selection-background-color: #2563eb;
                selection-color: white;
            }
            QCalendarWidget QToolButton {
                color: #0f172a;
                background-color: white;
            }
            QCalendarWidget QMenu {
                color: #0f172a;
                background-color: white;
            }
            QCalendarWidget QSpinBox {
                color: #0f172a;
                background-color: white;
            }
        """)
        filters_layout.addWidget(self.date_from)
        
        filters_layout.addWidget(QLabel("Hasta:"))
        self.date_to = QDateEdit()
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setCalendarPopup(True)
        self.date_to.setDisplayFormat("dd/MM/yyyy")
        self.date_to.setMinimumHeight(35)
        self.date_to.setStyleSheet("""
            QDateEdit {
                color: #0f172a;
                background-color: white;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                padding: 5px 10px;
            }
            QCalendarWidget QWidget {
                color: #0f172a;
            }
            QCalendarWidget QAbstractItemView {
                color: #0f172a;
                background-color: white;
                selection-background-color: #2563eb;
                selection-color: white;
            }
            QCalendarWidget QToolButton {
                color: #0f172a;
                background-color: white;
            }
            QCalendarWidget QMenu {
                color: #0f172a;
                background-color: white;
            }
            QCalendarWidget QSpinBox {
                color: #0f172a;
                background-color: white;
            }
        """)
        filters_layout.addWidget(self.date_to)
        
        btn_export = QPushButton("Exportar a Excel")
        btn_export.setMinimumHeight(35)
        btn_export.setMinimumWidth(140)
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        btn_export.clicked.connect(self.export_to_excel)
        filters_layout.addWidget(btn_export)
        
        filters_layout.addStretch()
        filters_group.setLayout(filters_layout)
        content_layout.addWidget(filters_group)
        
        
        # Productos más vendidos
        top_products_group = QGroupBox("Top 10 Productos Más Vendidos")
        top_products_group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 15px;
                padding: 20px 15px 15px 15px;
                font-weight: bold;
                color: #0f172a;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                left: 10px;
                top: 5px;
                padding: 0 5px;
                background-color: white;
                color: #0f172a;
            }
        """)
        top_products_layout = QVBoxLayout()
        top_products_layout.setContentsMargins(10, 25, 10, 10)
        
        self.top_products_table = QTableWidget()
        self.top_products_table.setColumnCount(4)
        self.top_products_table.setHorizontalHeaderLabels([
            "Posición", "Producto", "Cantidad Vendida", "Total Vendido"
        ])
        header = self.top_products_table.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.top_products_table.setColumnWidth(0, 80)
        self.top_products_table.setColumnWidth(2, 150)
        self.top_products_table.setColumnWidth(3, 150)
        self.top_products_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.top_products_table.verticalHeader().setVisible(False)
        self.top_products_table.setMinimumHeight(250)  # Altura para mostrar ~5 filas
        
        top_products_layout.addWidget(self.top_products_table)
        top_products_group.setLayout(top_products_layout)
        content_layout.addWidget(top_products_group)
        
        # Top clientes
        top_customers_group = QGroupBox("Top 10 Mejores Clientes")
        top_customers_group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: #0f172a;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #0f172a;
            }
        """)
        top_customers_layout = QVBoxLayout()
        
        self.top_customers_table = QTableWidget()
        self.top_customers_table.setColumnCount(4)
        self.top_customers_table.setHorizontalHeaderLabels([
            "Posición", "Cliente", "N° Compras", "Total Comprado"
        ])
        header = self.top_customers_table.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.top_customers_table.setColumnWidth(0, 80)
        self.top_customers_table.setColumnWidth(2, 120)
        self.top_customers_table.setColumnWidth(3, 150)
        self.top_customers_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.top_customers_table.verticalHeader().setVisible(False)
        self.top_customers_table.setMinimumHeight(250)  # Altura para mostrar ~5 filas
        
        top_customers_layout.addWidget(self.top_customers_table)
        top_customers_group.setLayout(top_customers_layout)
        content_layout.addWidget(top_customers_group)
        
        # Productos con bajo stock
        low_stock_group = QGroupBox("Alertas de Stock Bajo")
        low_stock_group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: #0f172a;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #0f172a;
            }
        """)
        low_stock_layout = QVBoxLayout()
        
        self.low_stock_table = QTableWidget()
        self.low_stock_table.setColumnCount(4)
        self.low_stock_table.setHorizontalHeaderLabels([
            "Producto", "Stock Actual", "Stock Mínimo", "Estado"
        ])
        header = self.low_stock_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.low_stock_table.setColumnWidth(1, 120)
        self.low_stock_table.setColumnWidth(2, 120)
        self.low_stock_table.setColumnWidth(3, 120)
        self.low_stock_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.low_stock_table.verticalHeader().setVisible(False)
        self.low_stock_table.setMinimumHeight(250)  # Altura para mostrar ~5 filas
        
        low_stock_layout.addWidget(self.low_stock_table)
        low_stock_group.setLayout(low_stock_layout)
        content_layout.addWidget(low_stock_group)
        
        # Consumo de materias primas
        materials_consumption_group = QGroupBox("Consumo de Materias Primas (Período Seleccionado)")
        materials_consumption_group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: #0f172a;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #0f172a;
            }
        """)
        materials_consumption_layout = QVBoxLayout()
        
        self.materials_consumption_table = QTableWidget()
        self.materials_consumption_table.setColumnCount(5)
        self.materials_consumption_table.setHorizontalHeaderLabels([
            "Materia Prima", "Unidad", "Cantidad Consumida", "Costo Total", "Stock Actual"
        ])
        header = self.materials_consumption_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.materials_consumption_table.setColumnWidth(1, 80)
        self.materials_consumption_table.setColumnWidth(2, 150)
        self.materials_consumption_table.setColumnWidth(3, 130)
        self.materials_consumption_table.setColumnWidth(4, 120)
        self.materials_consumption_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.materials_consumption_table.verticalHeader().setVisible(False)
        self.materials_consumption_table.setMinimumHeight(250)
        
        materials_consumption_layout.addWidget(self.materials_consumption_table)
        materials_consumption_group.setLayout(materials_consumption_layout)
        content_layout.addWidget(materials_consumption_group)
        
        # Proyección de producción
        production_projection_group = QGroupBox("Proyección de Producción (Basada en Stock de Materias Primas)")
        production_projection_group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: #0f172a;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #0f172a;
            }
        """)
        production_projection_layout = QVBoxLayout()
        
        self.production_projection_table = QTableWidget()
        self.production_projection_table.setColumnCount(4)
        self.production_projection_table.setHorizontalHeaderLabels([
            "Producto", "Unidades Producibles", "Costo Real (Materias Primas)", "Precio Venta"
        ])
        header = self.production_projection_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.production_projection_table.setColumnWidth(1, 150)
        self.production_projection_table.setColumnWidth(2, 200)
        self.production_projection_table.setColumnWidth(3, 120)
        self.production_projection_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.production_projection_table.verticalHeader().setVisible(False)
        self.production_projection_table.setMinimumHeight(250)
        
        production_projection_layout.addWidget(self.production_projection_table)
        production_projection_group.setLayout(production_projection_layout)
        content_layout.addWidget(production_projection_group)
        
        # Materias primas con bajo stock
        low_materials_stock_group = QGroupBox("Materias Primas con Bajo Stock")
        low_materials_stock_group.setStyleSheet("""
            QGroupBox {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                font-weight: bold;
                color: #0f172a;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #0f172a;
            }
        """)
        low_materials_stock_layout = QVBoxLayout()
        
        self.low_materials_stock_table = QTableWidget()
        self.low_materials_stock_table.setColumnCount(5)
        self.low_materials_stock_table.setHorizontalHeaderLabels([
            "Materia Prima", "Unidad", "Stock Actual", "Stock Mínimo", "Estado"
        ])
        header = self.low_materials_stock_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.low_materials_stock_table.setColumnWidth(1, 80)
        self.low_materials_stock_table.setColumnWidth(2, 120)
        self.low_materials_stock_table.setColumnWidth(3, 120)
        self.low_materials_stock_table.setColumnWidth(4, 120)
        self.low_materials_stock_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.low_materials_stock_table.verticalHeader().setVisible(False)
        self.low_materials_stock_table.setMinimumHeight(250)
        
        low_materials_stock_layout.addWidget(self.low_materials_stock_table)
        low_materials_stock_group.setLayout(low_materials_stock_layout)
        content_layout.addWidget(low_materials_stock_group)
        
        content_layout.addStretch()
        
        scroll.setWidget(content_widget)
        layout.addWidget(scroll)
    
    def load_reports(self):
        """Carga todos los reportes"""
        self.load_top_products()
        self.load_top_customers()
        self.load_low_stock_products()
        self.load_materials_consumption()
        self.load_production_projection()
        self.load_low_materials_stock()
    
    def load_top_products(self):
        """Carga los productos más vendidos"""
        session = get_session()
        try:
            date_from = self.date_from.date().toPyDate()
            date_to = self.date_to.date().toPyDate()
            
            from datetime import datetime
            datetime_from = datetime.combine(date_from, datetime.min.time())
            datetime_to = datetime.combine(date_to, datetime.max.time())
            
            # Consulta de productos más vendidos (sin filtro de status)
            top_products = session.query(
                Product.name,
                func.sum(SaleItem.quantity).label('total_quantity'),
                func.sum(SaleItem.subtotal).label('total_amount')
            ).join(
                SaleItem, Product.id == SaleItem.product_id
            ).join(
                Sale, SaleItem.sale_id == Sale.id
            ).filter(
                Sale.created_at.between(datetime_from, datetime_to)
            ).group_by(
                Product.id, Product.name
            ).order_by(
                desc('total_quantity')
            ).limit(10).all()
            
            # Llenar tabla
            self.top_products_table.setRowCount(len(top_products))
            
            for row, (name, quantity, amount) in enumerate(top_products):
                self.top_products_table.setItem(row, 0, QTableWidgetItem(f"#{row + 1}"))
                self.top_products_table.setItem(row, 1, QTableWidgetItem(name))
                self.top_products_table.setItem(row, 2, QTableWidgetItem(f"{int(quantity)} unidades"))
                self.top_products_table.setItem(row, 3, QTableWidgetItem(f"${amount:,.2f}"))
            
        except Exception as e:
            print(f"Error al cargar productos más vendidos: {e}")
            import traceback
            traceback.print_exc()
        finally:
            close_session()
    
    def load_top_customers(self):
        """Carga los mejores clientes"""
        session = get_session()
        try:
            date_from = self.date_from.date().toPyDate()
            date_to = self.date_to.date().toPyDate()
            
            from datetime import datetime
            datetime_from = datetime.combine(date_from, datetime.min.time())
            datetime_to = datetime.combine(date_to, datetime.max.time())
            
            # Consulta de mejores clientes (sin filtro de status)
            top_customers = session.query(
                Customer.name,
                func.count(Sale.id).label('purchase_count'),
                func.sum(Sale.total).label('total_amount')
            ).join(
                Sale, Customer.id == Sale.customer_id
            ).filter(
                Sale.created_at.between(datetime_from, datetime_to)
            ).group_by(
                Customer.id, Customer.name
            ).order_by(
                desc('total_amount')
            ).limit(10).all()
            
            # Llenar tabla
            self.top_customers_table.setRowCount(len(top_customers))
            
            for row, (name, count, amount) in enumerate(top_customers):
                self.top_customers_table.setItem(row, 0, QTableWidgetItem(f"#{row + 1}"))
                self.top_customers_table.setItem(row, 1, QTableWidgetItem(name))
                self.top_customers_table.setItem(row, 2, QTableWidgetItem(f"{count} compras"))
                self.top_customers_table.setItem(row, 3, QTableWidgetItem(f"${amount:,.2f}"))
            
        except Exception as e:
            print(f"Error al cargar mejores clientes: {e}")
            import traceback
            traceback.print_exc()
        finally:
            close_session()
    
    def load_low_stock_products(self):
        """Carga productos con stock bajo"""
        session = get_session()
        try:
            # Productos con stock bajo o igual al mínimo
            low_stock_products = session.query(Product).filter(
                Product.stock <= Product.min_stock
            ).all()
            
            # Llenar tabla
            self.low_stock_table.setRowCount(len(low_stock_products))
            
            for row, product in enumerate(low_stock_products):
                self.low_stock_table.setItem(row, 0, QTableWidgetItem(product.name))
                
                stock_item = QTableWidgetItem(str(product.stock))
                stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.low_stock_table.setItem(row, 1, stock_item)
                
                min_stock_item = QTableWidgetItem(str(product.min_stock))
                min_stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.low_stock_table.setItem(row, 2, min_stock_item)
                
                # Estado
                if product.stock == 0:
                    status = "SIN STOCK"
                    color = Qt.GlobalColor.red
                elif product.stock <= product.min_stock:
                    status = "BAJO"
                    color = Qt.GlobalColor.darkYellow
                else:
                    status = "NORMAL"
                    color = Qt.GlobalColor.darkGreen
                
                status_item = QTableWidgetItem(status)
                status_item.setForeground(color)
                status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.low_stock_table.setItem(row, 3, status_item)
            
            if len(low_stock_products) == 0:
                self.low_stock_table.setRowCount(1)
                self.low_stock_table.setItem(0, 0, QTableWidgetItem("No hay productos con stock bajo"))
                self.low_stock_table.setSpan(0, 0, 1, 4)
            
        except Exception as e:
            print(f"Error al cargar productos con stock bajo: {e}")
        finally:
            close_session()
    
    def load_materials_consumption(self):
        """Carga el consumo de materias primas en el período seleccionado"""
        session = get_session()
        try:
            date_from = self.date_from.date().toPyDate()
            date_to = self.date_to.date().toPyDate()
            
            from datetime import datetime
            datetime_from = datetime.combine(date_from, datetime.min.time())
            datetime_to = datetime.combine(date_to, datetime.max.time())
            
            # Obtener movimientos de tipo PRODUCTION (ventas)
            movements = session.query(
                RawMaterialMovement.raw_material_id,
                func.sum(RawMaterialMovement.quantity).label('total_quantity')
            ).filter(
                RawMaterialMovement.movement_type == RawMaterialMovementType.PRODUCTION,
                RawMaterialMovement.created_at.between(datetime_from, datetime_to)
            ).group_by(RawMaterialMovement.raw_material_id).all()
            
            self.materials_consumption_table.setRowCount(len(movements))
            
            for row, (material_id, total_qty) in enumerate(movements):
                material = session.query(RawMaterial).filter_by(id=material_id).first()
                if not material:
                    continue
                
                # Nombre
                self.materials_consumption_table.setItem(row, 0, QTableWidgetItem(material.name))
                
                # Unidad
                unit_item = QTableWidgetItem(material.unit)
                unit_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.materials_consumption_table.setItem(row, 1, unit_item)
                
                # Cantidad consumida (absoluto)
                consumed = abs(total_qty)
                qty_item = QTableWidgetItem(f"{consumed:.2f}")
                qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.materials_consumption_table.setItem(row, 2, qty_item)
                
                # Costo total
                cost = consumed * material.cost_per_unit
                cost_item = QTableWidgetItem(f"${cost:,.2f}")
                cost_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
                self.materials_consumption_table.setItem(row, 3, cost_item)
                
                # Stock actual
                stock_item = QTableWidgetItem(f"{material.stock:.2f}")
                stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if material.is_low_stock:
                    stock_item.setForeground(Qt.GlobalColor.red)
                self.materials_consumption_table.setItem(row, 4, stock_item)
            
            if len(movements) == 0:
                self.materials_consumption_table.setRowCount(1)
                self.materials_consumption_table.setItem(0, 0, QTableWidgetItem("No hay consumo de materias primas en este período"))
                self.materials_consumption_table.setSpan(0, 0, 1, 5)
        
        except Exception as e:
            print(f"Error al cargar consumo de materias primas: {e}")
        finally:
            close_session()
    
    def load_production_projection(self):
        """Carga la proyección de producción basada en stock de materias primas"""
        session = get_session()
        try:
            # Obtener todos los productos con materias primas asociadas
            products = session.query(Product).options(joinedload(Product.product_materials)).all()
            
            # Filtrar productos que tienen materias primas
            products_with_materials = [p for p in products if p.product_materials]
            
            self.production_projection_table.setRowCount(len(products_with_materials))
            
            for row, product in enumerate(products_with_materials):
                # Nombre del producto
                self.production_projection_table.setItem(row, 0, QTableWidgetItem(product.name))
                
                # Unidades producibles
                max_units = product.max_producible_units
                units_item = QTableWidgetItem(str(max_units) if max_units != float('inf') else "∞")
                units_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if max_units == 0:
                    units_item.setForeground(Qt.GlobalColor.red)
                elif max_units < 10:
                    units_item.setForeground(Qt.GlobalColor.darkYellow)
                self.production_projection_table.setItem(row, 1, units_item)
                
                # Costo real de materias primas
                real_cost = product.real_cost_from_materials
                cost_item = QTableWidgetItem(f"${real_cost:.2f}")
                cost_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
                self.production_projection_table.setItem(row, 2, cost_item)
                
                # Precio de venta
                price_item = QTableWidgetItem(f"${product.sale_price:.2f}")
                price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
                self.production_projection_table.setItem(row, 3, price_item)
            
            if len(products_with_materials) == 0:
                self.production_projection_table.setRowCount(1)
                self.production_projection_table.setItem(0, 0, QTableWidgetItem("No hay productos con materias primas asignadas"))
                self.production_projection_table.setSpan(0, 0, 1, 4)
        
        except Exception as e:
            print(f"Error al cargar proyección de producción: {e}")
        finally:
            close_session()
    
    def load_low_materials_stock(self):
        """Carga las materias primas con stock bajo"""
        session = get_session()
        try:
            # Obtener materias primas con stock bajo o en cero
            materials = session.query(RawMaterial).filter(
                RawMaterial.stock <= RawMaterial.min_stock
            ).order_by(RawMaterial.stock.asc()).all()
            
            self.low_materials_stock_table.setRowCount(len(materials))
            
            for row, material in enumerate(materials):
                # Nombre
                name_item = QTableWidgetItem(material.name)
                if material.stock == 0:
                    name_item.setForeground(Qt.GlobalColor.red)
                self.low_materials_stock_table.setItem(row, 0, name_item)
                
                # Unidad
                unit_item = QTableWidgetItem(material.unit)
                unit_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.low_materials_stock_table.setItem(row, 1, unit_item)
                
                # Stock actual
                stock_item = QTableWidgetItem(f"{material.stock:.2f}")
                stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if material.stock == 0:
                    stock_item.setForeground(Qt.GlobalColor.red)
                self.low_materials_stock_table.setItem(row, 2, stock_item)
                
                # Stock mínimo
                min_stock_item = QTableWidgetItem(f"{material.min_stock:.2f}")
                min_stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.low_materials_stock_table.setItem(row, 3, min_stock_item)
                
                # Estado
                if material.stock == 0:
                    status = "SIN STOCK"
                    color = Qt.GlobalColor.red
                else:
                    status = "BAJO"
                    color = Qt.GlobalColor.darkYellow
                
                status_item = QTableWidgetItem(status)
                status_item.setForeground(color)
                status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.low_materials_stock_table.setItem(row, 4, status_item)
            
            if len(materials) == 0:
                self.low_materials_stock_table.setRowCount(1)
                self.low_materials_stock_table.setItem(0, 0, QTableWidgetItem("No hay materias primas con stock bajo"))
                self.low_materials_stock_table.setSpan(0, 0, 1, 5)
        
        except Exception as e:
            print(f"Error al cargar materias primas con stock bajo: {e}")
        finally:
            close_session()
    
    def export_to_excel(self):
        """Exporta el reporte a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            from PyQt6.QtWidgets import QFileDialog
            
            # Preguntar dónde guardar
            file_name, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Reporte",
                f"Reporte_Ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_name:
                return
            
            # Crear workbook
            wb = Workbook()
            
            # Hoja 1: Resumen
            ws1 = wb.active
            ws1.title = "Resumen"
            
            # Título
            ws1['A1'] = "REPORTE DE VENTAS"
            ws1['A1'].font = Font(size=16, bold=True)
            ws1['A2'] = f"Período: {self.date_from.date().toString('dd/MM/yyyy')} - {self.date_to.date().toString('dd/MM/yyyy')}"
            
            # Resumen de ventas (calculado dinámicamente)
            ws1['A4'] = "RESUMEN DE VENTAS"
            ws1['A4'].font = Font(bold=True)
            
            # Calcular resumen dinámicamente
            session = get_session()
            try:
                date_from = self.date_from.date().toPyDate()
                date_to = self.date_to.date().toPyDate()
                from datetime import datetime
                datetime_from = datetime.combine(date_from, datetime.min.time())
                datetime_to = datetime.combine(date_to, datetime.max.time())
                
                total_sales = session.query(func.sum(Sale.total)).filter(
                    Sale.created_at.between(datetime_from, datetime_to)
                ).scalar() or 0
                
                sales_count = session.query(Sale).filter(
                    Sale.created_at.between(datetime_from, datetime_to)
                ).count()
                
                avg_sale = total_sales / sales_count if sales_count > 0 else 0
                
                ws1['A5'] = "Total Ventas:"
                ws1['B5'] = f"${total_sales:,.2f}"
                ws1['A6'] = "Número de Ventas:"
                ws1['B6'] = str(sales_count)
                ws1['A7'] = "Ticket Promedio:"
                ws1['B7'] = f"${avg_sale:,.2f}"
            except Exception as e:
                ws1['A5'] = "Error al calcular resumen"
                ws1['B5'] = str(e)
            finally:
                close_session()
            
            # Hoja 2: Top Productos
            ws2 = wb.create_sheet("Top Productos")
            ws2['A1'] = "TOP 10 PRODUCTOS MÁS VENDIDOS"
            ws2['A1'].font = Font(size=14, bold=True)
            
            headers = ['Posición', 'Producto', 'Cantidad Vendida', 'Total Vendido']
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(3, col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row in range(self.top_products_table.rowCount()):
                for col in range(4):
                    item = self.top_products_table.item(row, col)
                    if item:
                        ws2.cell(row + 4, col + 1).value = item.text()
            
            # Hoja 3: Top Clientes
            ws3 = wb.create_sheet("Top Clientes")
            ws3['A1'] = "TOP 10 MEJORES CLIENTES"
            ws3['A1'].font = Font(size=14, bold=True)
            
            headers = ['Posición', 'Cliente', 'N° Compras', 'Total Comprado']
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(3, col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row in range(self.top_customers_table.rowCount()):
                for col in range(4):
                    item = self.top_customers_table.item(row, col)
                    if item:
                        ws3.cell(row + 4, col + 1).value = item.text()
            
            # Hoja 4: Stock Bajo
            ws4 = wb.create_sheet("Stock Bajo")
            ws4['A1'] = "ALERTAS DE STOCK BAJO"
            ws4['A1'].font = Font(size=14, bold=True)
            
            headers = ['Producto', 'Stock Actual', 'Stock Mínimo', 'Estado']
            for col, header in enumerate(headers, 1):
                cell = ws4.cell(3, col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row in range(self.low_stock_table.rowCount()):
                for col in range(4):
                    item = self.low_stock_table.item(row, col)
                    if item:
                        ws4.cell(row + 4, col + 1).value = item.text()
            
            # Ajustar ancho de columnas
            for ws in [ws1, ws2, ws3, ws4]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar
            wb.save(file_name)
            
            QMessageBox.information(
                self,
                "Éxito",
                f"Reporte exportado correctamente a:\n{file_name}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error al exportar a Excel:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()