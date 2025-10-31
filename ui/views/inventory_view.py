"""
Vista de Inventario - Control y movimientos de inventario
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QLineEdit, QMessageBox,
    QDialog, QFormLayout, QSpinBox, QComboBox, QHeaderView,
    QTextEdit, QDateEdit, QGroupBox, QTabWidget, QTimeEdit, QFileDialog
)
from PyQt6.QtCore import Qt, QDate, QDateTime
from PyQt6.QtGui import QShowEvent
from datetime import datetime, date
from sqlalchemy import update, func
from config.database import get_session, close_session
from models import Product, InventoryMovement, MovementType, RawMaterial, RawMaterialMovement, RawMaterialMovementType

class InventoryView(QWidget):
    """Vista para control de inventario"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
        self.load_inventory()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Header
        header_layout = QHBoxLayout()
        
        title = QLabel("Control de Inventario")
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #0f172a;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Botón para configurar contraseña
        btn_password = QPushButton("🔒 Contraseña")
        btn_password.setFixedHeight(40)
        btn_password.setToolTip("Configurar o cambiar contraseña de acceso")
        btn_password.setStyleSheet("""
            QPushButton {
                background-color: #8b5cf6;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #7c3aed;
            }
        """)
        btn_password.clicked.connect(self.manage_password)
        header_layout.addWidget(btn_password)
        
        # Botón para exportar movimientos del día a Excel
        btn_export = QPushButton("📊 Exportar a Excel")
        btn_export.setFixedHeight(40)
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        btn_export.clicked.connect(self.export_movements_to_excel)
        header_layout.addWidget(btn_export)
        
        # Botón para ajustar inventario
        btn_adjust = QPushButton("Ajustar Inventario")
        btn_adjust.setFixedHeight(40)
        btn_adjust.clicked.connect(self.adjust_inventory)
        header_layout.addWidget(btn_adjust)
        
        # Botón para resetear stock a cero
        btn_reset_stock = QPushButton("⚠️ Resetear Stock a Cero")
        btn_reset_stock.setFixedHeight(40)
        btn_reset_stock.setStyleSheet("""
            QPushButton {
                background-color: #ef4444;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #dc2626;
            }
        """)
        btn_reset_stock.clicked.connect(self.reset_all_stock_to_zero)
        header_layout.addWidget(btn_reset_stock)
        
        layout.addLayout(header_layout)
        
        # Tabs para Materia Prima y Productos
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                background-color: white;
            }
            QTabBar::tab {
                background-color: #f1f5f9;
                color: #64748b;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-size: 14px;
                font-weight: 500;
            }
            QTabBar::tab:selected {
                background-color: white;
                color: #0f172a;
                border-bottom: 2px solid #2563eb;
            }
            QTabBar::tab:hover {
                background-color: #e2e8f0;
            }
        """)
        
        # Tab de Materia Prima (primera)
        raw_materials_tab = QWidget()
        raw_materials_layout = QVBoxLayout(raw_materials_tab)
        raw_materials_layout.setContentsMargins(0, 0, 0, 0)
        
        # Barra de búsqueda para materia prima
        search_rm_layout = QHBoxLayout()
        self.search_rm_input = QLineEdit()
        self.search_rm_input.setPlaceholderText("Buscar materia prima...")
        self.search_rm_input.textChanged.connect(self.search_raw_materials)
        self.search_rm_input.setFixedHeight(40)
        self.search_rm_input.setStyleSheet("color: #0f172a;")
        search_rm_layout.addWidget(self.search_rm_input)
        
        btn_refresh_rm = QPushButton("Actualizar")
        btn_refresh_rm.setFixedWidth(120)
        btn_refresh_rm.setFixedHeight(40)
        btn_refresh_rm.clicked.connect(self.load_raw_materials_inventory)
        search_rm_layout.addWidget(btn_refresh_rm)
        raw_materials_layout.addLayout(search_rm_layout)
        
        # Tabla de materia prima
        self.raw_materials_table = QTableWidget()
        self.raw_materials_table.setColumnCount(5)
        self.raw_materials_table.setHorizontalHeaderLabels([
            "SKU", "Nombre", "Unidad", "Stock", "Stock Mín."
        ])
        
        header_rm = self.raw_materials_table.horizontalHeader()
        self.raw_materials_table.setColumnWidth(2, 100)  # Unidad
        self.raw_materials_table.setColumnWidth(3, 100)  # Stock
        self.raw_materials_table.setColumnWidth(4, 100)  # Stock Mín.
        
        header_rm.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # SKU
        header_rm.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Nombre
        
        self.raw_materials_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.raw_materials_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.raw_materials_table.verticalHeader().setVisible(False)
        self.raw_materials_table.horizontalHeader().setStretchLastSection(False)
        
        raw_materials_layout.addWidget(self.raw_materials_table)
        self.tabs.addTab(raw_materials_tab, "Materia Prima")
        
        # Tab de Productos (segunda)
        products_tab = QWidget()
        products_layout = QVBoxLayout(products_tab)
        products_layout.setContentsMargins(0, 0, 0, 0)
        
        # Barra de búsqueda para productos
        search_products_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar producto...")
        self.search_input.textChanged.connect(self.search_products)
        self.search_input.setFixedHeight(40)
        self.search_input.setStyleSheet("color: #0f172a;")
        search_products_layout.addWidget(self.search_input)
        
        btn_refresh = QPushButton("Actualizar")
        btn_refresh.setFixedWidth(120)
        btn_refresh.setFixedHeight(40)
        btn_refresh.clicked.connect(self.load_products_inventory)
        search_products_layout.addWidget(btn_refresh)
        products_layout.addLayout(search_products_layout)
        
        # Tabla de productos
        self.stock_table = QTableWidget()
        self.stock_table.setColumnCount(6)
        self.stock_table.setHorizontalHeaderLabels([
            "SKU", "Producto", "Categoría", "Stock", "Stock Mín.", "Estado"
        ])
        
        header = self.stock_table.horizontalHeader()
        self.stock_table.setColumnWidth(3, 80)   # Stock
        self.stock_table.setColumnWidth(4, 100)  # Stock Mín.
        self.stock_table.setColumnWidth(5, 100)  # Estado
        
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # SKU
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Producto
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # Categoría
        
        self.stock_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.stock_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.stock_table.verticalHeader().setVisible(False)
        self.stock_table.horizontalHeader().setStretchLastSection(False)
        
        products_layout.addWidget(self.stock_table)
        self.tabs.addTab(products_tab, "Productos")
        
        layout.addWidget(self.tabs)
        
        # Filtro de fecha para movimientos
        filter_date_layout = QHBoxLayout()
        filter_date_label = QLabel("Filtrar movimientos por fecha:")
        filter_date_layout.addWidget(filter_date_label)
        
        self.date_filter = QDateEdit()
        self.date_filter.setDate(QDate.currentDate())
        self.date_filter.setCalendarPopup(True)
        self.date_filter.setFixedHeight(35)
        self.date_filter.setStyleSheet("""
            QDateEdit {
                color: #0f172a;
                background-color: white;
            }
        """)
        self.date_filter.dateChanged.connect(self.on_tab_changed)
        filter_date_layout.addWidget(self.date_filter)
        
        filter_date_layout.addStretch()
        
        # Historial de movimientos
        movements_group = QGroupBox("Historial de Movimientos del Día")
        movements_layout = QVBoxLayout()
        movements_layout.addLayout(filter_date_layout)
        
        self.movements_table = QTableWidget()
        self.movements_table.setColumnCount(8)
        self.movements_table.setHorizontalHeaderLabels([
            "Fecha", "Item", "Tipo", "Cantidad", "Razón", "Stock Resultante", "Anotación", "Acciones"
        ])
        
        # Conectar cambio de tab para actualizar movimientos
        self.tabs.currentChanged.connect(self.on_tab_changed)
        
        header = self.movements_table.horizontalHeader()
        # Configurar anchos fijos para columnas específicas
        self.movements_table.setColumnWidth(3, 80)   # Cantidad
        self.movements_table.setColumnWidth(5, 130)  # Stock Resultante
        
        # Hacer que varias columnas se expandan proporcionalmente
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # Fecha
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Producto
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # Tipo
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # Razón
        
        self.movements_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.movements_table.verticalHeader().setVisible(False)
        self.movements_table.setMaximumHeight(300)
        
        # Asegurar que la tabla use todo el espacio disponible
        self.movements_table.horizontalHeader().setStretchLastSection(False)
        
        movements_layout.addWidget(self.movements_table)
        movements_group.setLayout(movements_layout)
        layout.addWidget(movements_group)
    
    def showEvent(self, event: QShowEvent):
        """Se ejecuta cuando la vista se muestra"""
        super().showEvent(event)
        # Recargar inventario cada vez que se muestra la vista
        self.load_inventory()
    
    def load_inventory(self):
        """Carga el inventario actual (ambos tabs)"""
        self.load_raw_materials_inventory()
        self.load_products_inventory()
        self.on_tab_changed()
    
    def on_tab_changed(self):
        """Se llama cuando cambia el tab activo"""
        current_index = self.tabs.currentIndex()
        if current_index == 0:  # Materia Prima
            self.load_raw_materials_movements()
        else:  # Productos
            self.load_products_movements()
    
    def load_raw_materials_inventory(self):
        """Carga el inventario de materias primas"""
        session = get_session()
        try:
            materials = session.query(RawMaterial).all()
            
            self.raw_materials_table.setRowCount(len(materials))
            
            for row, material in enumerate(materials):
                # SKU
                self.raw_materials_table.setItem(row, 0, QTableWidgetItem(material.sku))
                
                # Nombre
                name_item = QTableWidgetItem(material.name)
                if material.is_low_stock:
                    name_item.setForeground(Qt.GlobalColor.red)
                self.raw_materials_table.setItem(row, 1, name_item)
                
                # Unidad
                self.raw_materials_table.setItem(row, 2, QTableWidgetItem(material.unit))
                
                # Stock
                stock_item = QTableWidgetItem(f"{material.stock:.2f}")
                stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if material.is_low_stock:
                    stock_item.setForeground(Qt.GlobalColor.red)
                self.raw_materials_table.setItem(row, 3, stock_item)
                
                # Stock mínimo
                min_stock_item = QTableWidgetItem(f"{material.min_stock:.2f}")
                min_stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.raw_materials_table.setItem(row, 4, min_stock_item)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar materia prima: {str(e)}")
        finally:
            close_session()
    
    def load_products_inventory(self):
        """Carga el inventario de productos"""
        session = get_session()
        try:
            products = session.query(Product).all()
            
            self.stock_table.setRowCount(len(products))
            
            for row, product in enumerate(products):
                # SKU
                self.stock_table.setItem(row, 0, QTableWidgetItem(product.sku))
                
                # Nombre
                self.stock_table.setItem(row, 1, QTableWidgetItem(product.name))
                
                # Categoría
                category = product.category.name if product.category else "-"
                self.stock_table.setItem(row, 2, QTableWidgetItem(category))
                
                # Stock
                stock_item = QTableWidgetItem(str(product.stock))
                stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.stock_table.setItem(row, 3, stock_item)
                
                # Stock mínimo
                min_stock_item = QTableWidgetItem(str(product.min_stock))
                min_stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.stock_table.setItem(row, 4, min_stock_item)
                
                # Estado
                if product.stock == 0:
                    status = "SIN STOCK"
                    color = Qt.GlobalColor.red
                elif product.is_low_stock:
                    status = "BAJO"
                    color = Qt.GlobalColor.darkYellow
                else:
                    status = "NORMAL"
                    color = Qt.GlobalColor.darkGreen
                
                status_item = QTableWidgetItem(status)
                status_item.setForeground(color)
                status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.stock_table.setItem(row, 5, status_item)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar productos: {str(e)}")
        finally:
            close_session()
    
    def search_raw_materials(self, text):
        """Busca materias primas en la tabla"""
        for row in range(self.raw_materials_table.rowCount()):
            sku = self.raw_materials_table.item(row, 0).text().lower()
            name = self.raw_materials_table.item(row, 1).text().lower()
            
            if text.lower() in sku or text.lower() in name:
                self.raw_materials_table.setRowHidden(row, False)
            else:
                self.raw_materials_table.setRowHidden(row, True)
    
    def load_products_movements(self):
        """Carga el historial de movimientos de productos"""
        session = get_session()
        try:
            # Obtener la fecha seleccionada
            selected_date = self.date_filter.date().toPyDate()
            
            # Consultar movimientos del día seleccionado
            movements = session.query(InventoryMovement).filter(
                func.date(InventoryMovement.created_at) == selected_date
            ).order_by(InventoryMovement.created_at.desc()).all()
            
            self.movements_table.setRowCount(len(movements))
            
            for row, movement in enumerate(movements):
                # Verificar si está anulado ANTES de crear los items
                note_text = getattr(movement, 'note', None) or "-"
                is_cancelled = note_text.startswith("[ANULADO]")
                
                # Fecha
                date_str = movement.created_at.strftime("%d/%m/%Y %H:%M")
                date_item = QTableWidgetItem(date_str)
                if is_cancelled:
                    date_item.setBackground(Qt.GlobalColor.lightGray)
                    date_item.setForeground(Qt.GlobalColor.gray)
                self.movements_table.setItem(row, 0, date_item)
                
                # Producto
                product_name = movement.product.name if movement.product else "Producto eliminado"
                product_item = QTableWidgetItem(product_name)
                if is_cancelled:
                    product_item.setBackground(Qt.GlobalColor.lightGray)
                    product_item.setForeground(Qt.GlobalColor.gray)
                self.movements_table.setItem(row, 1, product_item)
                
                # Tipo
                type_item = QTableWidgetItem(movement.movement_type.value)
                if is_cancelled:
                    type_item.setBackground(Qt.GlobalColor.lightGray)
                    type_item.setForeground(Qt.GlobalColor.gray)
                elif movement.movement_type == MovementType.ENTRY:
                    type_item.setForeground(Qt.GlobalColor.darkGreen)
                elif movement.movement_type == MovementType.EXIT:
                    type_item.setForeground(Qt.GlobalColor.red)
                self.movements_table.setItem(row, 2, type_item)
                
                # Cantidad
                quantity_item = QTableWidgetItem(str(abs(movement.quantity)))
                quantity_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if is_cancelled:
                    quantity_item.setBackground(Qt.GlobalColor.lightGray)
                    quantity_item.setForeground(Qt.GlobalColor.gray)
                self.movements_table.setItem(row, 3, quantity_item)
                
                # Razón
                reason_text = movement.reason or "-"
                reason_item = QTableWidgetItem(reason_text)
                if is_cancelled:
                    reason_item.setBackground(Qt.GlobalColor.lightGray)
                    reason_item.setForeground(Qt.GlobalColor.gray)
                self.movements_table.setItem(row, 4, reason_item)
                
                # Stock resultante
                stock_item = QTableWidgetItem(str(movement.new_stock))
                stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if is_cancelled:
                    stock_item.setBackground(Qt.GlobalColor.lightGray)
                    stock_item.setForeground(Qt.GlobalColor.gray)
                self.movements_table.setItem(row, 5, stock_item)
                
                # Nota (columna separada)
                note_item = QTableWidgetItem(note_text)
                if is_cancelled:
                    note_item.setForeground(Qt.GlobalColor.red)
                    note_item.setBackground(Qt.GlobalColor.lightGray)
                elif note_text != "-":
                    note_item.setForeground(Qt.GlobalColor.blue)
                self.movements_table.setItem(row, 6, note_item)
                
                # Botón para agregar anotación
                btn_note = QPushButton("📝")
                btn_note.setFixedSize(36, 30)
                btn_note.setToolTip("Agregar anotación")
                btn_note.setStyleSheet("""
                    QPushButton {
                        background-color: #8b5cf6;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 14px;
                        padding: 0px;
                    }
                    QPushButton:hover {
                        background-color: #7c3aed;
                    }
                """)
                btn_note.clicked.connect(lambda checked, m=movement: self.add_note_to_movement(m))
                
                # Contenedor para centrar el botón
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.addWidget(btn_note)
                actions_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                
                self.movements_table.setCellWidget(row, 7, actions_widget)
                
                # Aumentar altura de la fila para que el botón se vea bien
                self.movements_table.setRowHeight(row, 50)
            
        except Exception as e:
            print(f"Error al cargar movimientos de productos: {e}")
        finally:
            close_session()
    
    def manage_password(self):
        """Gestiona la contraseña de acceso al inventario"""
        from config.database import get_session, close_session
        from models.inventory_password import InventoryPassword
        
        session = get_session()
        try:
            password_record = session.query(InventoryPassword).first()
            has_password = password_record and password_record.password_hash
        finally:
            close_session()
        
        if has_password:
            # Cambiar contraseña
            from ui.dialogs.inventory_auth_dialog import ChangeInventoryPasswordDialog
            dialog = ChangeInventoryPasswordDialog(self)
            dialog.exec()
        else:
            # Establecer contraseña
            from ui.dialogs.inventory_auth_dialog import InventoryAuthDialog
            dialog = InventoryAuthDialog(self, is_password_set=False)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                QMessageBox.information(
                    self,
                    "✅ Contraseña Configurada",
                    "La sección de inventario ahora está protegida con contraseña."
                )
    
    def load_raw_materials_movements(self):
        """Carga el historial de movimientos de materia prima"""
        session = get_session()
        try:
            # Obtener la fecha seleccionada
            selected_date = self.date_filter.date().toPyDate()
            
            # Consultar movimientos del día seleccionado
            movements = session.query(RawMaterialMovement).filter(
                func.date(RawMaterialMovement.created_at) == selected_date
            ).order_by(RawMaterialMovement.created_at.desc()).all()
            
            self.movements_table.setRowCount(len(movements))
            
            for row, movement in enumerate(movements):
                # Fecha
                date_str = movement.created_at.strftime("%d/%m/%Y %H:%M")
                self.movements_table.setItem(row, 0, QTableWidgetItem(date_str))
                
                # Materia prima
                material_name = movement.raw_material.name if movement.raw_material else "Materia prima eliminada"
                self.movements_table.setItem(row, 1, QTableWidgetItem(material_name))
                
                # Tipo
                type_item = QTableWidgetItem(movement.movement_type.value)
                if movement.quantity > 0:  # Entrada
                    type_item.setForeground(Qt.GlobalColor.darkGreen)
                else:  # Salida
                    type_item.setForeground(Qt.GlobalColor.red)
                self.movements_table.setItem(row, 2, type_item)
                
                # Cantidad
                quantity_item = QTableWidgetItem(f"{abs(movement.quantity):.2f}")
                quantity_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.movements_table.setItem(row, 3, quantity_item)
                
                # Razón
                reason_text = movement.reason or "-"
                reason_item = QTableWidgetItem(reason_text)
                self.movements_table.setItem(row, 4, reason_item)
                
                # Stock resultante (obtener stock actual del material)
                current_stock = movement.raw_material.stock if movement.raw_material else 0
                stock_item = QTableWidgetItem(f"{current_stock:.2f}")
                stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.movements_table.setItem(row, 5, stock_item)
                
                # Nota (columna separada)
                note_text = getattr(movement, 'note', None) or "-"
                note_item = QTableWidgetItem(note_text)
                if note_text != "-":
                    note_item.setForeground(Qt.GlobalColor.blue)
                self.movements_table.setItem(row, 6, note_item)
                
                # Botón para agregar anotación
                btn_note = QPushButton("📝")
                btn_note.setFixedSize(36, 30)
                btn_note.setToolTip("Agregar anotación")
                btn_note.setStyleSheet("""
                    QPushButton {
                        background-color: #8b5cf6;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 14px;
                        padding: 0px;
                    }
                    QPushButton:hover {
                        background-color: #7c3aed;
                    }
                """)
                btn_note.clicked.connect(lambda checked, m=movement: self.add_note_to_movement(m, is_raw_material=True))
                
                # Contenedor para centrar el botón
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(2, 2, 2, 2)
                actions_layout.addWidget(btn_note)
                actions_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
                
                self.movements_table.setCellWidget(row, 7, actions_widget)
                self.movements_table.setRowHeight(row, 50)
            
        except Exception as e:
            print(f"Error al cargar movimientos de materia prima: {e}")
        finally:
            close_session()
    
    def search_products(self, text):
        """Busca productos en la tabla"""
        for row in range(self.stock_table.rowCount()):
            sku = self.stock_table.item(row, 0).text().lower()
            name = self.stock_table.item(row, 1).text().lower()
            
            if text.lower() in sku or text.lower() in name:
                self.stock_table.setRowHidden(row, False)
            else:
                self.stock_table.setRowHidden(row, True)
    
    def add_note_to_movement(self, movement, is_raw_material=False):
        """Abre diálogo para agregar una anotación a un movimiento"""
        dialog = AddNoteDialog(self, movement, is_raw_material)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.on_tab_changed()  # Recargar movimientos según el tab activo
    
    def export_movements_to_excel(self):
        """Exporta los movimientos del día seleccionado a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.styles.colors import Color
            import os
            
            # Solicitar ubicación para guardar
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar movimientos del día",
                os.path.join(os.path.expanduser("~"), "Desktop", f"Movimientos_{self.date_filter.date().toString('yyyyMMdd')}.xlsx"),
                "Archivos Excel (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Obtener la fecha seleccionada
            selected_date = self.date_filter.date().toPyDate()
            
            session = get_session()
            try:
                # Obtener movimientos de productos del día
                product_movements = session.query(InventoryMovement).filter(
                    func.date(InventoryMovement.created_at) == selected_date
                ).order_by(InventoryMovement.created_at.asc()).all()
                
                # Obtener movimientos de materias primas del día
                raw_material_movements = session.query(RawMaterialMovement).filter(
                    func.date(RawMaterialMovement.created_at) == selected_date
                ).order_by(RawMaterialMovement.created_at.asc()).all()
                
                # Crear libro de Excel
                wb = Workbook()
                wb.remove(wb.active)  # Eliminar hoja por defecto
                
                # Hoja 1: Movimientos de Productos
                ws_products = wb.create_sheet(title="Productos")
                
                # Título
                ws_products.merge_cells('A1:H1')
                title_cell = ws_products['A1']
                title_cell.value = f"Movimientos de Productos - {selected_date.strftime('%d/%m/%Y')}"
                title_cell.font = Font(size=14, bold=True, color="FFFFFF")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
                
                # Encabezados
                headers = ["Fecha/Hora", "Producto", "Tipo", "Cantidad", "Stock Anterior", "Stock Nuevo", "Razón", "Anotación"]
                for col, header in enumerate(headers, 1):
                    cell = ws_products.cell(row=3, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Datos de productos
                for row, movement in enumerate(product_movements, 4):
                    ws_products.cell(row=row, column=1, value=movement.created_at.strftime("%d/%m/%Y %H:%M:%S"))
                    ws_products.cell(row=row, column=2, value=movement.product.name if movement.product else "Producto eliminado")
                    ws_products.cell(row=row, column=3, value=movement.movement_type.value)
                    ws_products.cell(row=row, column=4, value=movement.quantity)
                    ws_products.cell(row=row, column=5, value=movement.previous_stock)
                    ws_products.cell(row=row, column=6, value=movement.new_stock)
                    ws_products.cell(row=row, column=7, value=movement.reason or "")
                    ws_products.cell(row=row, column=8, value=getattr(movement, 'note', '') or "")
                
                # Ajustar anchos de columna para productos
                ws_products.column_dimensions['A'].width = 20
                ws_products.column_dimensions['B'].width = 30
                ws_products.column_dimensions['C'].width = 12
                ws_products.column_dimensions['D'].width = 12
                ws_products.column_dimensions['E'].width = 15
                ws_products.column_dimensions['F'].width = 15
                ws_products.column_dimensions['G'].width = 40
                ws_products.column_dimensions['H'].width = 40
                
                # Hoja 2: Movimientos de Materias Primas
                ws_materials = wb.create_sheet(title="Materias Primas")
                
                # Título
                ws_materials.merge_cells('A1:G1')
                title_cell = ws_materials['A1']
                title_cell.value = f"Movimientos de Materias Primas - {selected_date.strftime('%d/%m/%Y')}"
                title_cell.font = Font(size=14, bold=True, color="FFFFFF")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                
                # Encabezados
                headers = ["Fecha/Hora", "Materia Prima", "Tipo", "Cantidad", "Unidad", "Razón", "Costo"]
                for col, header in enumerate(headers, 1):
                    cell = ws_materials.cell(row=3, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Datos de materias primas
                for row, movement in enumerate(raw_material_movements, 4):
                    material_name = movement.raw_material.name if movement.raw_material else "Materia prima eliminada"
                    material_unit = movement.raw_material.unit if movement.raw_material else ""
                    
                    ws_materials.cell(row=row, column=1, value=movement.created_at.strftime("%d/%m/%Y %H:%M:%S"))
                    ws_materials.cell(row=row, column=2, value=material_name)
                    ws_materials.cell(row=row, column=3, value=movement.movement_type.value)
                    ws_materials.cell(row=row, column=4, value=movement.quantity)
                    ws_materials.cell(row=row, column=5, value=material_unit)
                    ws_materials.cell(row=row, column=6, value=movement.reason or "")
                    ws_materials.cell(row=row, column=7, value=movement.cost or 0)
                
                # Ajustar anchos de columna para materias primas
                ws_materials.column_dimensions['A'].width = 20
                ws_materials.column_dimensions['B'].width = 30
                ws_materials.column_dimensions['C'].width = 15
                ws_materials.column_dimensions['D'].width = 12
                ws_materials.column_dimensions['E'].width = 12
                ws_materials.column_dimensions['F'].width = 40
                ws_materials.column_dimensions['G'].width = 12
                
                # Guardar archivo
                wb.save(file_path)
                
                total_productos = len(product_movements)
                total_materias = len(raw_material_movements)
                
                QMessageBox.information(
                    self,
                    "✅ Éxito",
                    f"Movimientos exportados correctamente.\n\n"
                    f"Fecha: {selected_date.strftime('%d/%m/%Y')}\n"
                    f"Productos: {total_productos} movimientos\n"
                    f"Materias Primas: {total_materias} movimientos\n"
                    f"Archivo: {file_path}"
                )
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}")
            finally:
                close_session()
                
        except ImportError:
            QMessageBox.warning(
                self,
                "Error",
                "La librería openpyxl no está instalada.\n\n"
                "Instale con: pip install openpyxl"
            )
    
    def adjust_inventory(self):
        """Abre diálogo para ajustar inventario"""
        current_index = self.tabs.currentIndex()
        is_raw_material = (current_index == 0)  # 0 = Materia Prima
        
        dialog = InventoryAdjustmentDialog(self, is_raw_material=is_raw_material)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventory()
    
    def reset_all_stock_to_zero(self):
        """Resetea el stock de todos los productos a cero"""
        # Confirmar acción
        reply = QMessageBox.question(
            self,
            "⚠️ Confirmar Acción",
            "¿Está seguro de que desea resetear el stock de TODOS los productos a cero?\n\n"
            "Esta acción NO se puede deshacer.\n"
            "Se registrará un movimiento de ajuste para cada producto.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Segunda confirmación adicional por seguridad
        reply2 = QMessageBox.question(
            self,
            "⚠️ Última Confirmación",
            "Esta acción reseteará el stock de TODOS los productos a 0.\n\n"
            "¿Está completamente seguro de continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply2 != QMessageBox.StandardButton.Yes:
            return
        
        session = get_session()
        try:
            products = session.query(Product).all()
            
            if not products:
                QMessageBox.information(self, "Info", "No hay productos en el inventario")
                return
            
            # Contar productos con stock > 0
            products_with_stock = [p for p in products if p.stock > 0]
            
            if not products_with_stock:
                QMessageBox.information(self, "Info", "Todos los productos ya tienen stock en cero")
                return
            
            # Resetear stock de cada producto con stock > 0
            reset_count = 0
            for product in products_with_stock:
                previous_stock = product.stock
                
                # Calcular cantidad para el movimiento
                quantity_for_movement = -previous_stock  # Negativo porque es salida
                
                # Registrar movimiento de ajuste
                movement = InventoryMovement(
                    product_id=product.id,
                    movement_type=MovementType.ADJUSTMENT,
                    quantity=quantity_for_movement,
                    previous_stock=previous_stock,
                    new_stock=0,
                    reason=f"Reset masivo de inventario - Stock anterior: {previous_stock}"
                )
                
                # Actualizar stock del producto
                product.stock = 0
                
                session.add(movement)
                reset_count += 1
            
            session.commit()
            
            QMessageBox.information(
                self,
                "✅ Éxito",
                f"Stock reseteado correctamente.\n\n"
                f"Productos afectados: {reset_count}\n"
                f"Todos los productos ahora tienen stock en cero."
            )
            
            # Recargar inventario
            self.load_inventory()
            
        except Exception as e:
            session.rollback()
            QMessageBox.critical(
                self,
                "Error",
                f"Error al resetear stock:\n{str(e)}"
            )
        finally:
            close_session()


class AddNoteDialog(QDialog):
    """Diálogo para agregar una anotación a un movimiento"""
    def __init__(self, parent=None, movement=None, is_raw_material=False):
        super().__init__(parent)
        self.movement = movement
        self.is_raw_material = is_raw_material
        self.setWindowTitle("Agregar Anotación")
        self.setMinimumWidth(500)
        self.init_ui()
    
    def init_ui(self):
        self.setStyleSheet("""
            QDialog {
                background-color: white;
            }
            QLabel {
                color: #0f172a;
                font-size: 13px;
            }
            QTextEdit {
                color: #0f172a;
                background-color: white;
            }
        """)
        
        layout = QVBoxLayout(self)
        
        # Información del movimiento
        if self.is_raw_material:
            item_name = self.movement.raw_material.name if self.movement.raw_material else "Materia prima eliminada"
        else:
            item_name = self.movement.product.name if self.movement.product else "Producto eliminado"
        
        item_label = QLabel(f"Item: {item_name}")
        item_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #0f172a;")
        layout.addWidget(item_label)
        
        date_label = QLabel(f"Fecha: {self.movement.created_at.strftime('%d/%m/%Y %H:%M')}")
        date_label.setStyleSheet("font-size: 13px; color: #64748b;")
        layout.addWidget(date_label)
        
        # Anotación actual (si existe)
        current_note_label = QLabel("Anotación actual:")
        current_note_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #0f172a; margin-top: 10px;")
        layout.addWidget(current_note_label)
        
        current_note_text = getattr(self.movement, 'note', None)
        if not current_note_text:
            current_note_text = "-"
        current_note = QLabel(current_note_text)
        current_note.setStyleSheet("font-size: 12px; color: #64748b; padding: 8px; background-color: #f1f5f9; border-radius: 4px;")
        current_note.setWordWrap(True)
        layout.addWidget(current_note)
        
        # Nueva anotación
        new_note_label = QLabel("Nueva anotación:")
        new_note_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #0f172a; margin-top: 10px;")
        layout.addWidget(new_note_label)
        
        self.note_input = QTextEdit()
        self.note_input.setPlaceholderText("Escriba su anotación aquí...")
        self.note_input.setMinimumHeight(100)
        layout.addWidget(self.note_input)
        
        # Botones
        buttons_layout = QHBoxLayout()
        
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setMinimumHeight(40)
        btn_cancel.clicked.connect(self.reject)
        
        btn_save = QPushButton("Guardar Anotación")
        btn_save.setMinimumHeight(40)
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #8b5cf6;
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                padding: 8px 20px;
            }
            QPushButton:hover {
                background-color: #7c3aed;
            }
        """)
        btn_save.clicked.connect(self.save_note)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(btn_cancel)
        buttons_layout.addWidget(btn_save)
        
        layout.addLayout(buttons_layout)
    
    def save_note(self):
        """Guarda la anotación"""
        note = self.note_input.toPlainText().strip()
        
        if not note:
            QMessageBox.warning(self, "Error", "Debe escribir una anotación")
            return
        
        session = get_session()
        try:
            if self.is_raw_material:
                movement = session.query(RawMaterialMovement).filter_by(id=self.movement.id).first()
            else:
                movement = session.query(InventoryMovement).filter_by(id=self.movement.id).first()
            
            if not movement:
                QMessageBox.warning(self, "Error", "Movimiento no encontrado")
                return
            
            # Guardar anotación en el campo note
            movement.note = note
            
            session.commit()
            
            QMessageBox.information(
                self,
                "Éxito",
                "Anotación guardada correctamente"
            )
            self.accept()
            
        except Exception as e:
            session.rollback()
            QMessageBox.critical(self, "Error", f"Error al guardar anotación: {str(e)}")
        finally:
            close_session()


class InventoryAdjustmentDialog(QDialog):
    """Diálogo para ajustar inventario"""
    def __init__(self, parent=None, is_raw_material=False):
        super().__init__(parent)
        self.is_raw_material = is_raw_material
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("Ajustar Inventario")
        self.setMinimumWidth(500)
        self.setStyleSheet("""
            QDialog {
                background-color: white;
            }
            QLabel {
                color: #0f172a;
                font-size: 13px;
                background-color: transparent;
            }
            QLineEdit, QTextEdit, QComboBox, QSpinBox {
                color: #0f172a;
                background-color: white;
            }
        """)
        
        layout = QVBoxLayout(self)
        
        # Formulario
        form_layout = QFormLayout()
        
        # Producto o Materia Prima
        label_text = "Materia Prima*:" if self.is_raw_material else "Producto*:"
        self.product_combo = QComboBox()
        self.product_combo.setMinimumHeight(35)
        self.product_combo.setStyleSheet("""
            QComboBox {
                color: #0f172a;
                background-color: white;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #0f172a;
                selection-background-color: #3b82f6;
                selection-color: white;
            }
        """)
        self.load_products()
        self.product_combo.currentIndexChanged.connect(self.on_product_changed)
        form_layout.addRow(label_text, self.product_combo)
        
        # Stock actual
        self.current_stock_label = QLabel("0")
        self.current_stock_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2563eb;")
        form_layout.addRow("Stock Actual:", self.current_stock_label)
        
        # Tipo de movimiento
        self.movement_type_combo = QComboBox()
        self.movement_type_combo.setMinimumHeight(35)
        self.movement_type_combo.addItem("Entrada (+)", MovementType.ENTRY)
        self.movement_type_combo.addItem("Salida (-)", MovementType.EXIT)
        self.movement_type_combo.addItem("Ajuste", MovementType.ADJUSTMENT)
        self.movement_type_combo.currentIndexChanged.connect(self.update_preview)
        form_layout.addRow("Tipo de Movimiento*:", self.movement_type_combo)
        
        # Cantidad
        self.quantity_spin = QSpinBox()
        self.quantity_spin.setMinimum(1)
        self.quantity_spin.setMaximum(99999)
        self.quantity_spin.setMinimumHeight(35)
        self.quantity_spin.setButtonSymbols(QSpinBox.ButtonSymbols.NoButtons)
        self.quantity_spin.valueChanged.connect(self.update_preview)
        form_layout.addRow("Cantidad*:", self.quantity_spin)
        
        # Stock resultante (preview)
        self.new_stock_label = QLabel("0")
        self.new_stock_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #10b981;")
        form_layout.addRow("Nuevo Stock:", self.new_stock_label)
        
        # Razón
        self.reason_input = QTextEdit()
        self.reason_input.setMaximumHeight(80)
        self.reason_input.setPlaceholderText("Motivo del ajuste (ej: Compra a proveedor, Inventario físico, etc.)")
        form_layout.addRow("Razón:", self.reason_input)
        
        layout.addLayout(form_layout)
        
        # Botones
        buttons_layout = QHBoxLayout()
        
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setMinimumHeight(40)
        btn_cancel.clicked.connect(self.reject)
        
        btn_save = QPushButton("Guardar Ajuste")
        btn_save.setMinimumHeight(40)
        btn_save.clicked.connect(self.save_adjustment)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(btn_cancel)
        buttons_layout.addWidget(btn_save)
        
        layout.addLayout(buttons_layout)
        
        # Inicializar
        self.on_product_changed()
    
    def load_products(self):
        """Carga los productos o materias primas según corresponda"""
        self.product_combo.clear()
        session = get_session()
        try:
            if self.is_raw_material:
                # Cargar materias primas
                materials = session.query(RawMaterial).all()
                for material in materials:
                    display_text = f"{material.name} (Stock: {material.stock:.2f} {material.unit})"
                    self.product_combo.addItem(display_text, material.id)
            else:
                # Cargar productos
                products = session.query(Product).all()
                for product in products:
                    display_text = f"{product.name} (Stock: {product.stock})"
                    self.product_combo.addItem(display_text, product.id)
        finally:
            close_session()
    
    def on_product_changed(self):
        """Actualiza el stock actual cuando cambia el producto o materia prima"""
        item_id = self.product_combo.currentData()
        if not item_id:
            return
        
        session = get_session()
        try:
            if self.is_raw_material:
                material = session.query(RawMaterial).filter_by(id=item_id).first()
                if material:
                    self.current_stock_label.setText(f"{material.stock:.2f} {material.unit}")
                    self.update_preview()
            else:
                product = session.query(Product).filter_by(id=item_id).first()
                if product:
                    self.current_stock_label.setText(str(product.stock))
                    self.update_preview()
        finally:
            close_session()
    
    def update_preview(self):
        """Actualiza la vista previa del nuevo stock"""
        try:
            stock_text = self.current_stock_label.text()
            if self.is_raw_material:
                # Para materia prima, extraer solo el número
                current_stock = float(stock_text.split()[0])
            else:
                current_stock = int(stock_text)
            
            quantity = self.quantity_spin.value()
            movement_type = self.movement_type_combo.currentData()
            
            if movement_type == MovementType.ENTRY:
                new_stock = current_stock + quantity
            elif movement_type == MovementType.EXIT:
                new_stock = current_stock - quantity
            else:  # ADJUSTMENT
                new_stock = quantity  # Para ajustes, la cantidad es el nuevo stock
            
            # Validar que no sea negativo
            if new_stock < 0:
                self.new_stock_label.setText("ERROR: Stock negativo")
                self.new_stock_label.setStyleSheet("font-size: 16px; font-weight: bold; color: red;")
            else:
                if self.is_raw_material:
                    unit = stock_text.split()[-1] if len(stock_text.split()) > 1 else ""
                    self.new_stock_label.setText(f"{new_stock:.2f} {unit}")
                else:
                    self.new_stock_label.setText(str(new_stock))
                self.new_stock_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #10b981;")
        except:
            pass
    
    def save_adjustment(self):
        """Guarda el ajuste de inventario"""
        item_id = self.product_combo.currentData()
        if not item_id:
            item_type = "materia prima" if self.is_raw_material else "producto"
            QMessageBox.warning(self, "Error", f"Seleccione una {item_type}")
            return
        
        reason = self.reason_input.toPlainText().strip()
        if not reason:
            reason = "Ajuste de inventario"  # Razón por defecto si está vacío
        
        # Validar stock negativo
        if "ERROR" in self.new_stock_label.text():
            QMessageBox.warning(self, "Error", "El ajuste resultaría en stock negativo")
            return
        
        session = get_session()
        try:
            if self.is_raw_material:
                # Ajustar materia prima
                material = session.query(RawMaterial).filter_by(id=item_id).first()
                if not material:
                    QMessageBox.warning(self, "Error", "Materia prima no encontrada")
                    return
                
                previous_stock = material.stock
                quantity = self.quantity_spin.value()
                movement_type = self.movement_type_combo.currentData()
                
                # Calcular nuevo stock
                if movement_type == MovementType.ENTRY:
                    material.stock += quantity
                    quantity_for_movement = quantity
                elif movement_type == MovementType.EXIT:
                    material.stock -= quantity
                    quantity_for_movement = -quantity
                else:  # ADJUSTMENT
                    material.stock = quantity
                    quantity_for_movement = quantity - previous_stock
                
                # Mapear MovementType a RawMaterialMovementType
                from models import RawMaterialMovementType
                if movement_type == MovementType.ENTRY:
                    rm_type = RawMaterialMovementType.PURCHASE
                elif movement_type == MovementType.EXIT:
                    rm_type = RawMaterialMovementType.PRODUCTION
                else:
                    rm_type = RawMaterialMovementType.ADJUSTMENT
                
                # Registrar movimiento
                movement = RawMaterialMovement(
                    raw_material_id=material.id,
                    movement_type=rm_type,
                    quantity=quantity_for_movement,
                    reason=reason
                )
                
                session.add(movement)
                session.commit()
                
                QMessageBox.information(
                    self,
                    "Éxito",
                    f"Inventario ajustado correctamente\n"
                    f"Stock anterior: {previous_stock:.2f} {material.unit}\n"
                    f"Nuevo stock: {material.stock:.2f} {material.unit}"
                )
                
            else:
                # Ajustar producto
                product = session.query(Product).filter_by(id=item_id).first()
                if not product:
                    QMessageBox.warning(self, "Error", "Producto no encontrado")
                    return
                
                previous_stock = product.stock
                quantity = self.quantity_spin.value()
                movement_type = self.movement_type_combo.currentData()
                
                # Calcular nuevo stock
                if movement_type == MovementType.ENTRY:
                    product.stock += quantity
                    quantity_for_movement = quantity
                elif movement_type == MovementType.EXIT:
                    product.stock -= quantity
                    quantity_for_movement = -quantity
                else:  # ADJUSTMENT
                    product.stock = quantity
                    quantity_for_movement = quantity - previous_stock
                
                # Registrar movimiento
                movement = InventoryMovement(
                    product_id=product.id,
                    movement_type=movement_type,
                    quantity=quantity_for_movement,
                    previous_stock=previous_stock,
                    new_stock=product.stock,
                    reason=reason
                )
                
                session.add(movement)
                session.commit()
                
                QMessageBox.information(
                    self,
                    "Éxito",
                    f"Inventario ajustado correctamente\n"
                    f"Stock anterior: {previous_stock}\n"
                    f"Nuevo stock: {product.stock}"
                )
            
            self.accept()
            
        except Exception as e:
            session.rollback()
            QMessageBox.critical(self, "Error", f"Error al guardar ajuste: {str(e)}")
        finally:
            close_session()