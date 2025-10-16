"""
Vista de Clientes - Gestión completa de clientes
"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QLineEdit, QMessageBox,
    QDialog, QFormLayout, QComboBox, QTextEdit, QHeaderView, QFileDialog
)
from PyQt6.QtCore import Qt
from PyQt6.QtGui import QShowEvent
from config.database import get_session, close_session
from models import Customer
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import traceback

class CustomersView(QWidget):
    """Vista para gestionar clientes"""
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.load_customers()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Header
        header_layout = QHBoxLayout()
        
        title = QLabel("Gestión de Clientes")
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #0f172a;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Botón para limpiar tabla
        btn_clear = QPushButton("🗑️ Limpiar Tabla")
        btn_clear.setFixedHeight(40)
        btn_clear.setFixedWidth(140)
        btn_clear.setStyleSheet("""
            QPushButton {
                background-color: #ef4444;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #dc2626;
            }
        """)
        btn_clear.clicked.connect(self.clear_customers_table)
        header_layout.addWidget(btn_clear)
        
        # Botón para importar Excel
        btn_import = QPushButton("📥 Importar Excel")
        btn_import.setFixedHeight(40)
        btn_import.setFixedWidth(150)
        btn_import.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        btn_import.clicked.connect(self.import_from_excel)
        header_layout.addWidget(btn_import)
        
        # Botón para exportar Excel
        btn_export = QPushButton("📤 Exportar Excel")
        btn_export.setFixedHeight(40)
        btn_export.setFixedWidth(150)
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        btn_export.clicked.connect(self.export_to_excel)
        header_layout.addWidget(btn_export)
        
        # Botón para agregar cliente
        btn_add = QPushButton("+ Nuevo Cliente")
        btn_add.setFixedHeight(40)
        btn_add.clicked.connect(self.add_customer)
        header_layout.addWidget(btn_add)
        
        layout.addLayout(header_layout)
        
        # Barra de búsqueda
        search_layout = QHBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar por nombre, email o documento...")
        self.search_input.textChanged.connect(self.search_customers)
        self.search_input.setFixedHeight(40)
        self.search_input.setStyleSheet("color: #0f172a;")
        search_layout.addWidget(self.search_input)
        
        btn_refresh = QPushButton("Actualizar")
        btn_refresh.setFixedWidth(120)
        btn_refresh.setFixedHeight(40)
        btn_refresh.clicked.connect(self.load_customers)
        search_layout.addWidget(btn_refresh)
        
        layout.addLayout(search_layout)
        
        # Tabla de clientes
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "ID", "Nombre", "Email", "Teléfono", "Documento", "Acciones"
        ])
        
        # Configurar tabla
        header = self.table.horizontalHeader()
        # Configurar anchos fijos para columnas específicas
        self.table.setColumnWidth(0, 50)   # ID
        self.table.setColumnWidth(5, 220)  # Acciones
        
        # Hacer que varias columnas se expandan proporcionalmente
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Nombre
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # Email
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # Teléfono
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # Documento
        
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        
        # Asegurar que la tabla use todo el espacio disponible
        self.table.horizontalHeader().setStretchLastSection(False)
        
        layout.addWidget(self.table)
    
    def showEvent(self, event: QShowEvent):
        """Se ejecuta cuando la vista se muestra"""
        super().showEvent(event)
        self.load_customers()
    
    def load_customers(self):
        """Carga todos los clientes en la tabla"""
        session = get_session()
        try:
            customers = session.query(Customer).all()
            
            self.table.setRowCount(len(customers))
            
            for row, customer in enumerate(customers):
                # ID
                self.table.setItem(row, 0, QTableWidgetItem(str(customer.id)))
                
                # Nombre
                self.table.setItem(row, 1, QTableWidgetItem(customer.name))
                
                # Email
                email = customer.email or "-"
                self.table.setItem(row, 2, QTableWidgetItem(email))
                
                # Teléfono
                phone = customer.phone or "-"
                self.table.setItem(row, 3, QTableWidgetItem(phone))
                
                # Documento
                document = customer.document_number or "-"
                self.table.setItem(row, 4, QTableWidgetItem(document))
                
                # Botones de acción
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(8, 5, 8, 5)
                actions_layout.setSpacing(8)
                
                btn_edit = QPushButton("Editar")
                btn_edit.setMinimumWidth(80)
                btn_edit.setFixedHeight(32)
                btn_edit.setStyleSheet("""
                    QPushButton {
                        background-color: #2563eb;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 12px;
                        font-weight: 500;
                        padding: 4px 8px;
                    }
                    QPushButton:hover {
                        background-color: #1d4ed8;
                    }
                """)
                btn_edit.clicked.connect(lambda checked, c=customer: self.edit_customer(c))
                
                btn_delete = QPushButton("Eliminar")
                btn_delete.setMinimumWidth(80)
                btn_delete.setFixedHeight(32)
                btn_delete.setStyleSheet("""
                    QPushButton {
                        background-color: #ef4444;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        font-size: 12px;
                        font-weight: 500;
                        padding: 4px 8px;
                    }
                    QPushButton:hover {
                        background-color: #dc2626;
                    }
                """)
                btn_delete.clicked.connect(lambda checked, c=customer: self.delete_customer(c))
                
                actions_layout.addWidget(btn_edit)
                actions_layout.addWidget(btn_delete)
                
                self.table.setCellWidget(row, 5, actions_widget)
                self.table.setRowHeight(row, 50)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar clientes: {str(e)}")
        finally:
            close_session()
    
    def search_customers(self, text):
        """Busca clientes por nombre, email o documento"""
        for row in range(self.table.rowCount()):
            name = self.table.item(row, 1).text().lower()
            email = self.table.item(row, 2).text().lower()
            document = self.table.item(row, 4).text().lower()
            
            if (text.lower() in name or 
                text.lower() in email or 
                text.lower() in document):
                self.table.setRowHidden(row, False)
            else:
                self.table.setRowHidden(row, True)
    
    def add_customer(self):
        """Abre diálogo para agregar cliente"""
        dialog = CustomerDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_customers()
    
    def edit_customer(self, customer):
        """Abre diálogo para editar cliente"""
        dialog = CustomerDialog(self, customer)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_customers()
    
    def delete_customer(self, customer):
        """Elimina un cliente"""
        # Verificar si tiene ventas asociadas
        session = get_session()
        try:
            from models import Sale
            sales_count = session.query(Sale).filter_by(customer_id=customer.id).count()
            
            if sales_count > 0:
                reply = QMessageBox.question(
                    self,
                    "Cliente con ventas",
                    f"Este cliente tiene {sales_count} venta(s) registrada(s).\n"
                    "Las ventas no se eliminarán. ¿Desea continuar?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                
                if reply == QMessageBox.StandardButton.No:
                    return
            else:
                reply = QMessageBox.question(
                    self,
                    "Confirmar eliminación",
                    f"¿Está seguro de eliminar al cliente '{customer.name}'?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                
                if reply == QMessageBox.StandardButton.No:
                    return
            
            session.delete(customer)
            session.commit()
            QMessageBox.information(self, "Éxito", "Cliente eliminado correctamente")
            self.load_customers()
            
        except Exception as e:
            session.rollback()
            QMessageBox.critical(self, "Error", f"Error al eliminar: {str(e)}")
        finally:
            close_session()
    
    def export_to_excel(self):
        """Exporta todos los clientes a un archivo Excel"""
        try:
            # Seleccionar ubicación del archivo
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Clientes como Excel",
                f"Clientes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            session = get_session()
            try:
                # Obtener todos los clientes
                customers = session.query(Customer).order_by(Customer.name).all()
                
                if not customers:
                    QMessageBox.warning(self, "Advertencia", "No hay clientes para exportar")
                    return
                
                # Crear libro de Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Clientes"
                
                # Encabezados con estilo
                headers = ["ID", "Nombre", "Email", "Teléfono", "Dirección", "Tipo Doc.", "Número Doc.", "Fecha Registro"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Datos
                for row, customer in enumerate(customers, 2):
                    ws.cell(row=row, column=1, value=customer.id)
                    ws.cell(row=row, column=2, value=customer.name)
                    ws.cell(row=row, column=3, value=customer.email or "")
                    ws.cell(row=row, column=4, value=customer.phone or "")
                    ws.cell(row=row, column=5, value=customer.address or "")
                    ws.cell(row=row, column=6, value=customer.document_type or "")
                    ws.cell(row=row, column=7, value=customer.document_number or "")
                    ws.cell(row=row, column=8, value=customer.created_at.strftime("%d/%m/%Y %H:%M") if customer.created_at else "")
                
                # Ajustar anchos de columna
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 30
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 40
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 15
                ws.column_dimensions['H'].width = 18
                
                # Guardar archivo
                wb.save(file_path)
                
                QMessageBox.information(
                    self,
                    "Éxito",
                    f"Clientes exportados correctamente\n\nArchivo: {file_path}\nTotal clientes: {len(customers)}"
                )
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}\n\n{traceback.format_exc()}")
            finally:
                close_session()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")
    
    def import_from_excel(self):
        """Importa clientes desde un archivo Excel"""
        try:
            # Seleccionar archivo
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar archivo Excel",
                "",
                "Excel Files (*.xlsx *.xls)"
            )
            
            if not file_path:
                return
            
            # Mostrar mensaje de información sobre el formato esperado
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("Formato de Excel para Importación")
            msg.setText("El archivo Excel debe tener las siguientes columnas:")
            msg.setInformativeText(
                "Columnas esperadas:\n"
                "• ID (se generará automáticamente, se puede omitir)\n"
                "• Nombre (obligatorio)\n"
                "• Email (opcional)\n"
                "• Teléfono (opcional)\n"
                "• Dirección (opcional)\n"
                "• Tipo Documento (opcional: DNI, RUC, etc.)\n"
                "• Número Documento (opcional)\n\n"
                "NOTA: Se omitirá la primera fila (encabezados).\n"
                "Los clientes con nombres o emails duplicados serán omitidos.\n\n"
                "¿Desea continuar con la importación?"
            )
            msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            if msg.exec() != QMessageBox.StandardButton.Yes:
                return
            
            session = get_session()
            try:
                # Cargar archivo Excel
                wb = load_workbook(file_path)
                ws = wb.active
                
                # Validar encabezados para asegurarse de que es un archivo de clientes
                headers = [cell.value for cell in ws[1]]
                
                # Verificar que tenga encabezados relacionados con clientes
                expected_headers_keywords = ['nombre', 'email', 'telefono', 'teléfono', 'direccion', 'dirección', 'documento']
                has_valid_headers = any(
                    any(keyword in str(header).lower() for keyword in expected_headers_keywords)
                    for header in headers if header
                )
                
                # Verificar que NO tenga encabezados de ventas
                invalid_headers_keywords = ['factura', 'venta', 'método pago', 'metodo pago', 'subtotal', 'impuesto', 'descuento']
                has_invalid_headers = any(
                    any(keyword in str(header).lower() for keyword in invalid_headers_keywords)
                    for header in headers if header
                )
                
                if has_invalid_headers or not has_valid_headers:
                    QMessageBox.critical(
                        self,
                        "Formato Incorrecto",
                        "El archivo Excel no tiene el formato correcto para importar clientes.\n\n"
                        "Parece ser un archivo de ventas u otro tipo de datos.\n\n"
                        "Por favor, seleccione un archivo Excel con el formato de clientes:\n"
                        "• Nombre\n"
                        "• Email\n"
                        "• Teléfono\n"
                        "• Dirección\n"
                        "• Tipo Documento\n"
                        "• Número Documento"
                    )
                    return
                
                imported_count = 0
                errors = []
                
                # Leer datos (empezando desde la fila 2, asumiendo que la 1 son encabezados)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        # Verificar que la fila no esté vacía
                        if not row or not row[0]:
                            continue
                        
                        # Detectar si la primera columna es ID numérico
                        if isinstance(row[0], int):
                            # El formato tiene ID en la primera columna (generado por exportación)
                            name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                            email = str(row[2]).strip() if len(row) > 2 and row[2] else None
                            phone = str(row[3]).strip() if len(row) > 3 and row[3] else None
                            address = str(row[4]).strip() if len(row) > 4 and row[4] else None
                            document_type = str(row[5]).strip() if len(row) > 5 and row[5] else None
                            document_number = str(row[6]).strip() if len(row) > 6 and row[6] else None
                        else:
                            # El formato no tiene ID, empieza con nombre
                            name = str(row[0]).strip() if row[0] else ""
                            email = str(row[1]).strip() if len(row) > 1 and row[1] else None
                            phone = str(row[2]).strip() if len(row) > 2 and row[2] else None
                            address = str(row[3]).strip() if len(row) > 3 and row[3] else None
                            document_type = str(row[4]).strip() if len(row) > 4 and row[4] else None
                            document_number = str(row[5]).strip() if len(row) > 5 and row[5] else None
                        
                        # Limpiar valores vacíos (convertir a None)
                        if email and not email.strip():
                            email = None
                        elif email:
                            email = email.strip()
                        
                        if phone and not phone.strip():
                            phone = None
                        elif phone:
                            phone = phone.strip()
                        
                        if address and not address.strip():
                            address = None
                        elif address:
                            address = address.strip()
                        
                        if document_type and not document_type.strip():
                            document_type = None
                        elif document_type:
                            document_type = document_type.strip()
                        
                        if document_number and not document_number.strip():
                            document_number = None
                        elif document_number:
                            document_number = document_number.strip()
                        
                        if not name or not name.strip():
                            errors.append(f"Fila {row_idx}: Nombre vacío")
                            continue
                        
                        name = name.strip()
                        
                        # Verificar si ya existe un cliente con ese nombre o email
                        existing = session.query(Customer).filter_by(name=name).first()
                        if existing:
                            errors.append(f"Fila {row_idx}: Cliente '{name}' ya existe")
                            continue
                        
                        if email:
                            existing_email = session.query(Customer).filter_by(email=email).first()
                            if existing_email:
                                errors.append(f"Fila {row_idx}: Email '{email}' ya existe")
                                continue
                        
                        if document_number:
                            existing_doc = session.query(Customer).filter_by(document_number=document_number).first()
                            if existing_doc:
                                errors.append(f"Fila {row_idx}: Número de documento '{document_number}' ya existe")
                                continue
                        
                        # Crear cliente
                        customer = Customer(
                            name=name,
                            email=email,
                            phone=phone,
                            address=address,
                            document_type=document_type,
                            document_number=document_number
                        )
                        
                        session.add(customer)
                        session.flush()  # Validar antes de continuar
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f"Fila {row_idx}: {str(e)}")
                        continue
                
                # Intentar hacer commit de todo
                try:
                    session.commit()
                except Exception as e:
                    session.rollback()
                    raise Exception(f"Error al guardar datos: {str(e)}")
                
                # Mostrar resultado
                result_msg = f"Importación completada\n\nClientes importados: {imported_count}"
                if errors:
                    result_msg += f"\n\nErrores encontrados:\n" + "\n".join(errors[:10])
                    if len(errors) > 10:
                        result_msg += f"\n... y {len(errors) - 10} errores más"
                
                QMessageBox.information(self, "Resultado de Importación", result_msg)
                self.load_customers()
                
            except Exception as e:
                session.rollback()
                QMessageBox.critical(self, "Error", f"Error al importar: {str(e)}\n\n{traceback.format_exc()}")
            finally:
                close_session()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")
    
    def clear_customers_table(self):
        """Limpia toda la tabla de clientes con confirmación"""
        try:
            # Primera confirmación
            reply = QMessageBox.question(
                self,
                "⚠️ Confirmar Limpieza",
                "¿Está seguro de que desea eliminar TODOS los clientes?\n\n"
                "Esta acción NO se puede deshacer y eliminará:\n"
                "• Todos los clientes registrados\n"
                "• ADVERTENCIA: Esto puede afectar las ventas existentes\n"
                "  que estén asociadas a estos clientes\n\n"
                "¿Desea continuar?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Segunda confirmación (seguridad adicional)
            reply2 = QMessageBox.warning(
                self,
                "⚠️ ÚLTIMA CONFIRMACIÓN",
                "Esta es su ÚLTIMA oportunidad para cancelar.\n\n"
                "¿Realmente desea eliminar TODOS los clientes de forma permanente?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply2 != QMessageBox.StandardButton.Yes:
                return
            
            session = get_session()
            try:
                # Verificar si hay ventas asociadas
                from models import Sale
                sales_count = session.query(Sale).filter(Sale.customer_id.isnot(None)).count()
                
                if sales_count > 0:
                    reply3 = QMessageBox.warning(
                        self,
                        "⚠️ ADVERTENCIA: Ventas Asociadas",
                        f"Hay {sales_count} ventas asociadas a clientes.\n\n"
                        "Al eliminar los clientes, estas ventas quedarán sin cliente asociado.\n\n"
                        "¿Desea continuar de todas formas?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )
                    
                    if reply3 != QMessageBox.StandardButton.Yes:
                        return
                
                # Contar clientes antes de eliminar
                total_customers = session.query(Customer).count()
                
                if total_customers == 0:
                    QMessageBox.information(self, "Información", "No hay clientes para eliminar")
                    return
                
                # Eliminar todos los clientes
                session.query(Customer).delete()
                
                session.commit()
                
                QMessageBox.information(
                    self,
                    "✅ Limpieza Completada",
                    f"Se han eliminado correctamente {total_customers} clientes"
                )
                
                self.load_customers()
                
            except Exception as e:
                session.rollback()
                QMessageBox.critical(self, "Error", f"Error al limpiar tabla: {str(e)}\n\n{traceback.format_exc()}")
            finally:
                close_session()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")


class CustomerDialog(QDialog):
    """Diálogo para crear/editar clientes"""
    def __init__(self, parent=None, customer=None):
        super().__init__(parent)
        self.customer = customer
        self.is_editing = customer is not None
        self.init_ui()
        
        if self.is_editing:
            self.load_customer_data()
    
    def init_ui(self):
        self.setWindowTitle("Editar Cliente" if self.is_editing else "Nuevo Cliente")
        self.setMinimumWidth(500)
        self.setStyleSheet("""
            QDialog {
                background-color: white;
            }
            QLabel {
                color: #0f172a;
                font-size: 13px;
                background-color: transparent;
            }
            QLineEdit, QTextEdit, QComboBox {
                color: #0f172a;
                background-color: white;
            }
        """)
        
        layout = QVBoxLayout(self)
        
        # Formulario
        form_layout = QFormLayout()
        
        # Nombre
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Nombre completo")
        self.name_input.setMinimumHeight(35)
        form_layout.addRow("Nombre*:", self.name_input)
        
        # Email
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("correo@ejemplo.com")
        self.email_input.setMinimumHeight(35)
        form_layout.addRow("Email:", self.email_input)
        
        # Teléfono
        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("Número de teléfono")
        self.phone_input.setMinimumHeight(35)
        form_layout.addRow("Teléfono:", self.phone_input)
        
        # Tipo de documento
        self.document_type_combo = QComboBox()
        self.document_type_combo.setMinimumHeight(35)
        self.document_type_combo.addItems(["DNI", "RUC", "Pasaporte", "Cédula", "Otro"])
        form_layout.addRow("Tipo Documento:", self.document_type_combo)
        
        # Número de documento
        self.document_number_input = QLineEdit()
        self.document_number_input.setPlaceholderText("Número de documento")
        self.document_number_input.setMinimumHeight(35)
        form_layout.addRow("N° Documento:", self.document_number_input)
        
        # Dirección
        self.address_input = QTextEdit()
        self.address_input.setMaximumHeight(80)
        self.address_input.setPlaceholderText("Dirección completa")
        form_layout.addRow("Dirección:", self.address_input)
        
        # Ciudad
        self.city_input = QLineEdit()
        self.city_input.setPlaceholderText("Ciudad")
        self.city_input.setMinimumHeight(35)
        form_layout.addRow("Ciudad:", self.city_input)
        
        layout.addLayout(form_layout)
        
        # Botones
        buttons_layout = QHBoxLayout()
        
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.setMinimumHeight(40)
        btn_cancel.clicked.connect(self.reject)
        
        btn_save = QPushButton("Guardar")
        btn_save.setMinimumHeight(40)
        btn_save.clicked.connect(self.save_customer)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(btn_cancel)
        buttons_layout.addWidget(btn_save)
        
        layout.addLayout(buttons_layout)
    
    def load_customer_data(self):
        """Carga los datos del cliente a editar"""
        self.name_input.setText(self.customer.name)
        self.email_input.setText(self.customer.email or "")
        self.phone_input.setText(self.customer.phone or "")
        
        if self.customer.document_type:
            index = self.document_type_combo.findText(self.customer.document_type)
            if index >= 0:
                self.document_type_combo.setCurrentIndex(index)
        
        self.document_number_input.setText(self.customer.document_number or "")
        self.address_input.setText(self.customer.address or "")
        self.city_input.setText(self.customer.city or "")
    
    def save_customer(self):
        """Guarda el cliente (crear o actualizar)"""
        # Validar campos requeridos
        name = self.name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "Campo requerido", "El nombre es obligatorio")
            return
        
        # Validar email si se proporcionó
        email = self.email_input.text().strip()
        if email and "@" not in email:
            QMessageBox.warning(self, "Email inválido", "Ingrese un email válido")
            return
        
        session = get_session()
        try:
            if self.is_editing:
                # Actualizar cliente existente
                customer = session.query(Customer).filter_by(id=self.customer.id).first()
            else:
                # Crear nuevo cliente
                customer = Customer()
            
            # Asignar valores
            customer.name = name
            customer.email = email if email else None
            customer.phone = self.phone_input.text().strip() or None
            customer.document_type = self.document_type_combo.currentText()
            customer.document_number = self.document_number_input.text().strip() or None
            customer.address = self.address_input.toPlainText().strip() or None
            customer.city = self.city_input.text().strip() or None
            
            if not self.is_editing:
                session.add(customer)
            
            session.commit()
            
            QMessageBox.information(
                self,
                "Éxito",
                "Cliente guardado correctamente"
            )
            self.accept()
            
        except Exception as e:
            session.rollback()
            QMessageBox.critical(self, "Error", f"Error al guardar: {str(e)}")
        finally:
            close_session()