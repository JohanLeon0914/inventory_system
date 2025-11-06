"""
Vista de Ventas - Sistema completo de registro de ventas
"""
import os
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QLineEdit, QMessageBox,
    QDialog, QFormLayout, QComboBox, QSpinBox, QDoubleSpinBox,
    QHeaderView, QTextEdit, QDateEdit, QGroupBox, QFileDialog, QScrollArea, QGridLayout, QApplication
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QShowEvent, QPixmap, QIcon, QFont
from datetime import datetime
from config.database import get_session, close_session
from models import Sale, SaleItem, Product, Customer, PaymentMethod, SaleStatus, InventoryMovement, MovementType, ProductMaterial, RawMaterial, RawMaterialMovement, RawMaterialMovementType
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import traceback

class SalesView(QWidget):
    """Vista para gestionar ventas"""
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.invoice_dialog = None
        self.load_sales()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Header
        header_layout = QHBoxLayout()
        
        title = QLabel("Registro de Ventas")
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #0f172a;")
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Botón para limpiar tabla
        btn_clear = QPushButton("🗑️ Limpiar Tabla")
        btn_clear.setFixedHeight(40)
        btn_clear.setFixedWidth(140)
        btn_clear.setStyleSheet("""
            QPushButton {
                background-color: #ef4444;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #dc2626;
            }
        """)
        btn_clear.clicked.connect(self.clear_sales_table)
        header_layout.addWidget(btn_clear)
        
        # Botón para importar Excel
        btn_import = QPushButton("📥 Importar Excel")
        btn_import.setFixedHeight(40)
        btn_import.setFixedWidth(150)
        btn_import.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #059669;
            }
        """)
        btn_import.clicked.connect(self.import_from_excel)
        header_layout.addWidget(btn_import)
        
        # Botón para exportar Excel
        btn_export = QPushButton("📤 Exportar Excel")
        btn_export.setFixedHeight(40)
        btn_export.setFixedWidth(150)
        btn_export.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
        """)
        btn_export.clicked.connect(self.export_to_excel)
        header_layout.addWidget(btn_export)
        
        # Botón para nueva venta
        btn_new_sale = QPushButton("+ Nueva Venta")
        btn_new_sale.setFixedHeight(40)
        btn_new_sale.clicked.connect(self.create_new_sale)
        header_layout.addWidget(btn_new_sale)
        
        # Botón para editar información de la empresa
        btn_company = QPushButton("🏢 Empresa")
        btn_company.setFixedHeight(40)
        btn_company.setFixedWidth(120)
        btn_company.setStyleSheet("""
            QPushButton {
                background-color: #8b5cf6;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 500;
                padding: 0 15px;
            }
            QPushButton:hover {
                background-color: #7c3aed;
            }
        """)
        btn_company.clicked.connect(self.edit_company_info)
        header_layout.addWidget(btn_company)
        
        layout.addLayout(header_layout)
        
        # Filtros
        filter_layout = QHBoxLayout()
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar por número de factura...")
        self.search_input.textChanged.connect(self.apply_sales_filters)
        self.search_input.setFixedHeight(40)
        self.search_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: #0f172a;
                border: 1px solid #d1d5db;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border-color: #3b82f6;
                outline: none;
            }
            QLineEdit::placeholder {
                color: #9ca3af;
            }
        """)
        filter_layout.addWidget(self.search_input)
        
        # Filtro por método de pago
        self.payment_filter = QComboBox()
        self.payment_filter.setMinimumHeight(40)
        self.payment_filter.addItems(["Todos", "Efectivo", "Tarjeta", "Transferencia"])
        self.payment_filter.setStyleSheet("""
            QComboBox {
                background-color: white;
                color: #0f172a;
                border: 1px solid #d1d5db;
                border-radius: 4px;
                padding: 8px;
                font-size: 14px;
            }
            QComboBox:focus {
                border-color: #3b82f6;
                outline: none;
            }
            QComboBox::drop-down {
                border: none;
                background-color: #f9fafb;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border: none;
                width: 8px;
                height: 8px;
                border-left: 4px solid transparent;
                border-right: 4px solid transparent;
                border-top: 6px solid #6b7280;
                margin-right: 8px;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #0f172a;
                border: 1px solid #d1d5db;
                selection-background-color: #3b82f6;
                selection-color: white;
            }
        """)
        self.payment_filter.currentIndexChanged.connect(self.apply_sales_filters)
        filter_layout.addWidget(self.payment_filter)
        
        btn_refresh = QPushButton("Actualizar")
        btn_refresh.setFixedWidth(120)
        btn_refresh.setFixedHeight(40)
        btn_refresh.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 4px;
                font-size: 14px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #2563eb;
            }
            QPushButton:pressed {
                background-color: #1d4ed8;
            }
        """)
        btn_refresh.clicked.connect(self.load_sales)
        filter_layout.addWidget(btn_refresh)
        
        # Total de ventas (visible y actualizado con los filtros)
        filter_layout.addStretch()
        self.sales_total_label = QLabel("Total Ventas: $0")
        self.sales_total_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #0f172a;")
        filter_layout.addWidget(self.sales_total_label)
        
        layout.addLayout(filter_layout)
        
        # Tabla de ventas
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels([
            "N° Factura", "Fecha", "Cliente", "Total", "Método Pago", "Estado", "Acciones"
        ])
        
        # Configurar tabla
        header = self.table.horizontalHeader()
        # Configurar anchos fijos para columnas específicas
        self.table.setColumnWidth(0, 120)  # N° Factura
        self.table.setColumnWidth(3, 120)  # Total
        self.table.setColumnWidth(6, 170)  # Acciones
        
        # Hacer que varias columnas se expandan proporcionalmente
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Fecha
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # Cliente
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # Método Pago
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)  # Estado
        
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.verticalHeader().setVisible(False)
        
        # Asegurar que la tabla use todo el espacio disponible
        self.table.horizontalHeader().setStretchLastSection(False)
        
        layout.addWidget(self.table)
    
    def showEvent(self, event: QShowEvent):
        """Se ejecuta cuando la vista se muestra"""
        super().showEvent(event)
        self.load_sales()
    
    def load_sales(self):
        """Carga todas las ventas"""
        session = get_session()
        try:
            sales = session.query(Sale).order_by(Sale.created_at.desc()).all()
            
            self.table.setRowCount(len(sales))
            
            for row, sale in enumerate(sales):
                # N° Factura
                self.table.setItem(row, 0, QTableWidgetItem(sale.invoice_number))
                
                # Fecha
                date_str = sale.created_at.strftime("%d/%m/%Y %H:%M")
                self.table.setItem(row, 1, QTableWidgetItem(date_str))
                
                # Cliente
                customer_name = sale.customer.name if sale.customer else "Cliente General"
                self.table.setItem(row, 2, QTableWidgetItem(customer_name))
                
                # Total
                if sale.total == int(sale.total):
                    total_text = f"${int(sale.total):,}"
                else:
                    total_text = f"${sale.total:,.2f}"
                total_item = QTableWidgetItem(total_text)
                total_item.setData(Qt.ItemDataRole.UserRole, float(sale.total))
                total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row, 3, total_item)
                
                # Método de pago
                self.table.setItem(row, 4, QTableWidgetItem(sale.payment_method.value))
                
                # Estado
                status_item = QTableWidgetItem(sale.status.value)
                if sale.status == SaleStatus.COMPLETED:
                    status_item.setForeground(Qt.GlobalColor.darkGreen)
                elif sale.status == SaleStatus.CANCELLED:
                    status_item.setForeground(Qt.GlobalColor.red)
                elif sale.status == SaleStatus.EDITED:
                    from PyQt6.QtGui import QColor
                    status_item.setForeground(QColor("#f59e0b"))  # Naranja para editada
                self.table.setItem(row, 5, status_item)
                
                # Botones de acción
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(4, 2, 4, 2)
                actions_layout.setSpacing(0)
                
                # Botón ver detalle (ícono 👁️, izquierda redondeada)
                btn_view = QPushButton("👁️")
                btn_view.setFixedSize(40, 32)
                btn_view.setToolTip("Ver Detalle")
                btn_view.setStyleSheet("""
                    QPushButton {
                        background-color: #2563eb;
                        color: white;
                        border: none;
                        border-top-left-radius: 4px;
                        border-bottom-left-radius: 4px;
                        border-top-right-radius: 0px;
                        border-bottom-right-radius: 0px;
                        font-size: 14px;
                        padding: 0px;
                    }
                    QPushButton:hover {
                        background-color: #1d4ed8;
                    }
                """)
                btn_view.clicked.connect(lambda checked, s=sale: self.view_sale_detail(s))
                actions_layout.addWidget(btn_view)
                
                # Botón editar (solo si no tiene factura) - sin bordes redondeados
                if not sale.has_invoice:
                    btn_edit = QPushButton("✏️")
                    btn_edit.setFixedSize(40, 32)
                    btn_edit.setToolTip("Editar")
                    btn_edit.setStyleSheet("""
                        QPushButton {
                            background-color: #f59e0b;
                            color: white;
                            border: none;
                            border-radius: 0px;
                            font-size: 14px;
                            padding: 0px;
                        }
                        QPushButton:hover {
                            background-color: #d97706;
                        }
                    """)
                    btn_edit.clicked.connect(lambda checked, s=sale: self.edit_sale_full(s))
                    actions_layout.addWidget(btn_edit)
                    
                    # Botón generar factura (verde, derecha redondeada)
                    btn_invoice = QPushButton("📄")
                    btn_invoice.setFixedSize(40, 32)
                    btn_invoice.setToolTip("Generar Factura")
                    btn_invoice.setStyleSheet("""
                        QPushButton {
                            background-color: #10b981;
                            color: white;
                            border: none;
                            border-top-left-radius: 0px;
                            border-bottom-left-radius: 0px;
                            border-top-right-radius: 4px;
                            border-bottom-right-radius: 4px;
                            font-size: 14px;
                            padding: 0px;
                        }
                        QPushButton:hover {
                            background-color: #059669;
                        }
                    """)
                    btn_invoice.clicked.connect(lambda checked, s=sale: self.generate_invoice(s))
                    actions_layout.addWidget(btn_invoice)
                else:
                    # Mostrar que ya tiene factura (verde, derecha redondeada)
                    btn_invoice_view = QPushButton("✅")
                    btn_invoice_view.setFixedSize(40, 32)
                    btn_invoice_view.setToolTip("Ver Factura")
                    btn_invoice_view.setStyleSheet("""
                        QPushButton {
                            background-color: #10b981;
                            color: white;
                            border: none;
                            border-top-left-radius: 0px;
                            border-bottom-left-radius: 0px;
                            border-top-right-radius: 4px;
                            border-bottom-right-radius: 4px;
                            font-size: 14px;
                            padding: 0px;
                        }
                        QPushButton:hover {
                            background-color: #059669;
                        }
                    """)
                    btn_invoice_view.clicked.connect(lambda checked, s=sale: self.view_invoice(s))
                    actions_layout.addWidget(btn_invoice_view)
                
                self.table.setCellWidget(row, 6, actions_widget)
                self.table.setRowHeight(row, 50)
            
            # Aplicar filtros vigentes y actualizar el total
            self.apply_sales_filters()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar ventas: {str(e)}")
        finally:
            close_session()

    def apply_sales_filters(self):
        """Aplica filtros por búsqueda y método de pago y actualiza el total."""
        text = self.search_input.text().lower() if hasattr(self, 'search_input') else ''
        method = self.payment_filter.currentText() if hasattr(self, 'payment_filter') else 'Todos'
        
        for row in range(self.table.rowCount()):
            invoice = (self.table.item(row, 0).text() if self.table.item(row, 0) else '').lower()
            pay_method = (self.table.item(row, 4).text() if self.table.item(row, 4) else '')
            
            match_text = text in invoice
            match_method = (method == 'Todos') or (pay_method == method)
            self.table.setRowHidden(row, not (match_text and match_method))
        
        self.update_sales_total_label()

    def update_sales_total_label(self):
        total = 0.0
        for row in range(self.table.rowCount()):
            if self.table.isRowHidden(row):
                continue
            
            # Verificar si la venta está cancelada
            status_item = self.table.item(row, 5)  # Columna "Estado"
            if status_item and status_item.text().lower() == "cancelada":
                continue  # Excluir ventas canceladas del total
            
            item = self.table.item(row, 3)
            if item:
                # Preferir valor crudo si existe
                value = item.data(Qt.ItemDataRole.UserRole)
                if value is None:
                    # Fallback parseando texto "$1,234.56"
                    try:
                        value = float(item.text().replace('$', '').replace(',', ''))
                    except:
                        value = 0.0
                total += float(value)
        
        # Formatear sin decimales si es entero
        if total == int(total):
            formatted_total = f"Total Ventas: ${int(total):,}"
        else:
            formatted_total = f"Total Ventas: ${total:,.2f}"
        
        self.sales_total_label.setText(formatted_total)
    
    def search_sales(self, text):
        """Busca ventas por número de factura"""
        for row in range(self.table.rowCount()):
            invoice = self.table.item(row, 0).text().lower()
            if text.lower() in invoice:
                self.table.setRowHidden(row, False)
            else:
                self.table.setRowHidden(row, True)
    
    def create_new_sale(self):
        """Abre el diálogo para crear una nueva venta"""
        from ui.views.new_sale_dialog import NewSaleDialog
        dialog = NewSaleDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_sales()
    
    def view_sale_detail(self, sale):
        """Muestra el detalle de una venta"""
        dialog = SaleDetailDialog(self, sale.id)
        dialog.exec()
    
    def edit_sale_full(self, sale):
        """Permite editar una venta completa incluyendo items y ajustando inventario"""
        dialog = EditSaleFullDialog(self, sale.id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_sales()
    
    def generate_invoice(self, sale):
        """Genera una factura legal colombiana para la venta"""
        from ui.views.invoice_dialog import InvoiceDialog
        
        # Confirmar generación
        reply = QMessageBox.question(
            self, "Confirmar Factura",
            f"¿Generar factura legal para la venta {sale.invoice_number}?\n\n"
            "Una vez generada, NO se podrá editar la venta.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.No:
            return
        
        # Abrir diálogo de factura
        dialog = InvoiceDialog(self, sale)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Marcar la venta como con factura
            session = get_session()
            try:
                updated_sale = session.query(Sale).filter_by(id=sale.id).first()
                if updated_sale:
                    updated_sale.has_invoice = 1
                    updated_sale.invoice_generated_at = datetime.now()
                    session.commit()
                    self.load_sales()
                    QMessageBox.information(self, "Éxito", "Factura generada correctamente.")
            except Exception as e:
                session.rollback()
                QMessageBox.critical(self, "Error", f"Error al marcar factura: {str(e)}")
            finally:
                close_session()
    
    def view_invoice(self, sale):
        """Muestra la factura generada"""
        from ui.views.invoice_dialog import InvoiceDialog
        dialog = InvoiceDialog(self, sale)
        dialog.exec()
    
    def edit_company_info(self):
        """Abre el diálogo para editar la información de la empresa"""
        from ui.views.company_info_dialog import CompanyInfoDialog
        dialog = CompanyInfoDialog(self)
        dialog.exec()
    
    def edit_sale_basic(self, sale):
        """Permite editar método de pago, estado, impuesto y descuento. No cambia items ni stock."""
        dialog = EditSaleDialog(self, sale)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_sales()
    
    def export_to_excel(self):
        """Exporta todas las ventas a un archivo Excel"""
        try:
            # Seleccionar ubicación del archivo
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Guardar Ventas como Excel",
                f"Ventas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            session = get_session()
            try:
                # Obtener todas las ventas con sus items
                from sqlalchemy.orm import joinedload
                sales = session.query(Sale).options(joinedload(Sale.items)).order_by(Sale.created_at.desc()).all()
                
                if not sales:
                    QMessageBox.warning(self, "Advertencia", "No hay ventas para exportar")
                    return
                
                # Crear libro de Excel
                wb = Workbook()
                
                # Hoja 1: Resumen de Ventas
                ws_summary = wb.active
                ws_summary.title = "Resumen de Ventas"
                
                # Encabezados con estilo
                headers = ["N° Factura", "Fecha", "Cliente", "Método Pago", "Estado", "Subtotal", "Impuesto", "Descuento", "Total"]
                for col, header in enumerate(headers, 1):
                    cell = ws_summary.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Datos
                for row, sale in enumerate(sales, 2):
                    ws_summary.cell(row=row, column=1, value=sale.invoice_number)
                    ws_summary.cell(row=row, column=2, value=sale.created_at.strftime("%d/%m/%Y %H:%M"))
                    ws_summary.cell(row=row, column=3, value=sale.customer.name if sale.customer else "Cliente General")
                    ws_summary.cell(row=row, column=4, value=sale.payment_method.value)
                    ws_summary.cell(row=row, column=5, value=sale.status.value)
                    ws_summary.cell(row=row, column=6, value=sale.subtotal)
                    ws_summary.cell(row=row, column=7, value=sale.tax)
                    ws_summary.cell(row=row, column=8, value=sale.discount)
                    ws_summary.cell(row=row, column=9, value=sale.total)
                
                # Ajustar anchos de columna
                ws_summary.column_dimensions['A'].width = 15
                ws_summary.column_dimensions['B'].width = 18
                ws_summary.column_dimensions['C'].width = 25
                ws_summary.column_dimensions['D'].width = 15
                ws_summary.column_dimensions['E'].width = 12
                ws_summary.column_dimensions['F'].width = 12
                ws_summary.column_dimensions['G'].width = 12
                ws_summary.column_dimensions['H'].width = 12
                ws_summary.column_dimensions['I'].width = 12
                
                # Hoja 2: Detalle de Items
                ws_detail = wb.create_sheet(title="Detalle de Items")
                
                detail_headers = ["N° Factura", "Fecha", "Cliente", "Producto", "Cantidad", "Precio Unit.", "Subtotal"]
                for col, header in enumerate(detail_headers, 1):
                    cell = ws_detail.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                detail_row = 2
                for sale in sales:
                    for item in sale.items:
                        ws_detail.cell(row=detail_row, column=1, value=sale.invoice_number)
                        ws_detail.cell(row=detail_row, column=2, value=sale.created_at.strftime("%d/%m/%Y %H:%M"))
                        ws_detail.cell(row=detail_row, column=3, value=sale.customer.name if sale.customer else "Cliente General")
                        ws_detail.cell(row=detail_row, column=4, value=item.product.name)
                        ws_detail.cell(row=detail_row, column=5, value=item.quantity)
                        ws_detail.cell(row=detail_row, column=6, value=item.unit_price)
                        ws_detail.cell(row=detail_row, column=7, value=item.subtotal)
                        detail_row += 1
                
                # Ajustar anchos de columna
                ws_detail.column_dimensions['A'].width = 15
                ws_detail.column_dimensions['B'].width = 18
                ws_detail.column_dimensions['C'].width = 25
                ws_detail.column_dimensions['D'].width = 30
                ws_detail.column_dimensions['E'].width = 10
                ws_detail.column_dimensions['F'].width = 12
                ws_detail.column_dimensions['G'].width = 12
                
                # Guardar archivo
                wb.save(file_path)
                
                QMessageBox.information(
                    self,
                    "Éxito",
                    f"Ventas exportadas correctamente\n\nArchivo: {file_path}\nTotal ventas: {len(sales)}"
                )
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al exportar: {str(e)}\n\n{traceback.format_exc()}")
            finally:
                close_session()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")
    
    def import_from_excel(self):
        """Importa ventas desde un archivo Excel"""
        try:
            # Seleccionar archivo
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar archivo Excel",
                "",
                "Excel Files (*.xlsx *.xls)"
            )
            
            if not file_path:
                return
            
            # Mostrar mensaje de información sobre el formato esperado
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Information)
            msg.setWindowTitle("Formato de Excel para Importación")
            msg.setText("El archivo Excel debe tener las siguientes columnas en la primera hoja:")
            msg.setInformativeText(
                "Columnas requeridas:\n"
                "• N° Factura (único, no debe existir)\n"
                "• Fecha (DD/MM/YYYY o DD/MM/YYYY HH:MM)\n"
                "• Cliente\n"
                "• Método Pago (Efectivo/Tarjeta/Transferencia)\n"
                "• Estado (Completada/Pendiente/Cancelada)\n"
                "• Subtotal (número)\n"
                "• Impuesto (número)\n"
                "• Descuento (número)\n"
                "• Total (número)\n\n"
                "NOTA: Las facturas duplicadas serán omitidas.\n\n"
                "¿Desea continuar con la importación?"
            )
            msg.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            if msg.exec() != QMessageBox.StandardButton.Yes:
                return
            
            session = get_session()
            try:
                # Cargar archivo Excel
                wb = load_workbook(file_path)
                ws = wb.active
                
                # Validar encabezados para asegurarse de que es un archivo de ventas
                headers = [cell.value for cell in ws[1]]
                
                # Verificar que tenga encabezados relacionados con ventas
                expected_headers_keywords = ['factura', 'fecha', 'cliente', 'método pago', 'metodo pago', 'estado', 'subtotal', 'total']
                has_valid_headers = any(
                    any(keyword in str(header).lower() for keyword in expected_headers_keywords)
                    for header in headers if header
                )
                
                # Verificar que NO tenga encabezados de clientes o productos
                invalid_headers_keywords = ['teléfono', 'telefono', 'dirección', 'direccion', 'stock', 'precio venta', 'precio costo']
                has_invalid_headers = any(
                    any(keyword in str(header).lower() for keyword in invalid_headers_keywords)
                    for header in headers if header
                )
                
                if has_invalid_headers or not has_valid_headers:
                    QMessageBox.critical(
                        self,
                        "Formato Incorrecto",
                        "El archivo Excel no tiene el formato correcto para importar ventas.\n\n"
                        "Parece ser un archivo de clientes, productos u otro tipo de datos.\n\n"
                        "Por favor, seleccione un archivo Excel con el formato de ventas:\n"
                        "• N° Factura\n"
                        "• Fecha\n"
                        "• Cliente\n"
                        "• Método Pago\n"
                        "• Estado\n"
                        "• Subtotal\n"
                        "• Impuesto\n"
                        "• Descuento\n"
                        "• Total"
                    )
                    return
                
                imported_count = 0
                errors = []
                
                # Leer datos (empezando desde la fila 2, asumiendo que la 1 son encabezados)
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        if not row[0]:  # Si no hay número de factura, saltar
                            continue
                        
                        invoice_number = str(row[0])
                        
                        # Verificar si ya existe la factura
                        existing = session.query(Sale).filter_by(invoice_number=invoice_number).first()
                        if existing:
                            errors.append(f"Fila {row_idx}: Factura {invoice_number} ya existe")
                            continue
                        
                        # Parsear fecha
                        try:
                            from datetime import date
                            if isinstance(row[1], datetime):
                                sale_date = row[1]
                            elif isinstance(row[1], date):
                                # Convertir date a datetime
                                sale_date = datetime.combine(row[1], datetime.min.time())
                            elif row[1]:
                                date_str = str(row[1]).strip()
                                # Intentar varios formatos de fecha
                                for date_format in ["%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"]:
                                    try:
                                        sale_date = datetime.strptime(date_str, date_format)
                                        break
                                    except:
                                        continue
                                else:
                                    raise ValueError("Formato no reconocido")
                            else:
                                raise ValueError("Fecha vacía")
                        except Exception as e:
                            errors.append(f"Fila {row_idx}: Formato de fecha inválido '{row[1]}' - {str(e)}")
                            continue
                        
                        # Buscar o crear cliente
                        customer_name = str(row[2]) if row[2] else "Cliente General"
                        customer = session.query(Customer).filter_by(name=customer_name).first()
                        if not customer and customer_name != "Cliente General":
                            customer = Customer(
                                name=customer_name,
                                email="",
                                phone="",
                                address=""
                            )
                            session.add(customer)
                            session.flush()
                        
                        # Método de pago
                        payment_method_str = str(row[3]) if row[3] else "Efectivo"
                        try:
                            payment_method = PaymentMethod(payment_method_str)
                        except:
                            payment_method = PaymentMethod.CASH
                        
                        # Estado
                        status_str = str(row[4]) if row[4] else "Completada"
                        try:
                            status = SaleStatus(status_str)
                        except:
                            status = SaleStatus.COMPLETED
                        
                        # Valores numéricos
                        try:
                            subtotal = float(row[5]) if row[5] else 0.0
                            tax = float(row[6]) if row[6] else 0.0
                            discount = float(row[7]) if row[7] else 0.0
                            total = float(row[8]) if row[8] else 0.0
                        except:
                            errors.append(f"Fila {row_idx}: Valores numéricos inválidos")
                            continue
                        
                        # Crear venta
                        sale = Sale(
                            invoice_number=invoice_number,
                            customer_id=customer.id if customer else None,
                            payment_method=payment_method,
                            status=status,
                            subtotal=subtotal,
                            tax=tax,
                            discount=discount,
                            total=total,
                            created_at=sale_date
                        )
                        
                        session.add(sale)
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f"Fila {row_idx}: {str(e)}")
                        continue
                
                session.commit()
                
                # Mostrar resultado
                result_msg = f"Importación completada\n\nVentas importadas: {imported_count}"
                if errors:
                    result_msg += f"\n\nErrores encontrados:\n" + "\n".join(errors[:10])
                    if len(errors) > 10:
                        result_msg += f"\n... y {len(errors) - 10} errores más"
                
                QMessageBox.information(self, "Resultado de Importación", result_msg)
                self.load_sales()
                
            except Exception as e:
                session.rollback()
                QMessageBox.critical(self, "Error", f"Error al importar: {str(e)}\n\n{traceback.format_exc()}")
            finally:
                close_session()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")
    
    def clear_sales_table(self):
        """Limpia toda la tabla de ventas con confirmación"""
        try:
            # Primera confirmación
            reply = QMessageBox.question(
                self,
                "⚠️ Confirmar Limpieza",
                "¿Está seguro de que desea eliminar TODAS las ventas?\n\n"
                "Esta acción NO se puede deshacer y eliminará:\n"
                "• Todas las ventas registradas\n"
                "• Todos los items de ventas\n"
                "• Los movimientos de inventario NO se eliminarán\n\n"
                "¿Desea continuar?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Segunda confirmación (seguridad adicional)
            reply2 = QMessageBox.warning(
                self,
                "⚠️ ÚLTIMA CONFIRMACIÓN",
                "Esta es su ÚLTIMA oportunidad para cancelar.\n\n"
                "¿Realmente desea eliminar TODAS las ventas de forma permanente?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply2 != QMessageBox.StandardButton.Yes:
                return
            
            session = get_session()
            try:
                # Contar ventas antes de eliminar
                total_sales = session.query(Sale).count()
                total_items = session.query(SaleItem).count()
                
                if total_sales == 0:
                    QMessageBox.information(self, "Información", "No hay ventas para eliminar")
                    return
                
                # Eliminar todos los items de ventas primero (por la relación)
                session.query(SaleItem).delete()
                
                # Eliminar todas las ventas
                session.query(Sale).delete()
                
                session.commit()
                
                QMessageBox.information(
                    self,
                    "✅ Limpieza Completada",
                    f"Se han eliminado correctamente:\n\n"
                    f"• {total_sales} ventas\n"
                    f"• {total_items} items de ventas"
                )
                
                self.load_sales()
                
            except Exception as e:
                session.rollback()
                QMessageBox.critical(self, "Error", f"Error al limpiar tabla: {str(e)}\n\n{traceback.format_exc()}")
            finally:
                close_session()
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error: {str(e)}")




class SaleDetailDialog(QDialog):
    """Diálogo para ver el detalle de una venta"""
    def __init__(self, parent=None, sale_id=None):
        super().__init__(parent)
        self.sale_id = sale_id
        self.sale_data = None
        self.items_data = []
        self.load_sale()
        self.init_ui()
    
    def load_sale(self):
        """Carga la venta desde la base de datos"""
        session = get_session()
        try:
            from sqlalchemy.orm import joinedload
            sale = session.query(Sale).options(joinedload(Sale.items)).filter_by(id=self.sale_id).first()
            if not sale:
                QMessageBox.warning(self, "Error", "Venta no encontrada")
                return
            
            # Extraer datos de la venta
            self.sale_data = {
                'invoice_number': sale.invoice_number,
                'created_at': sale.created_at,
                'customer_name': sale.customer.name if sale.customer else "Cliente General",
                'payment_method': sale.payment_method.value,
                'status': sale.status.value,
                'subtotal': sale.subtotal,
                'tax': sale.tax,
                'discount': sale.discount,
                'total': sale.total
            }
            
            # Extraer datos de los items
            self.items_data = []
            for item in sale.items:
                self.items_data.append({
                    'product_name': item.product.name,
                    'unit_price': item.unit_price,
                    'quantity': item.quantity,
                    'subtotal': item.subtotal
                })
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar venta: {str(e)}")
        finally:
            close_session()
    
    def init_ui(self):
        if not self.sale_data:
            self.reject()
            return
            
        self.setWindowTitle(f"Detalle de Venta - {self.sale_data['invoice_number']}")
        self.setMinimumSize(700, 500)
        self.setStyleSheet("""
            QDialog {
                background-color: white;
            }
            QLabel {
                color: #0f172a;
                background-color: transparent;
            }
            QTableWidget {
                color: #0f172a;
            }
            QTableWidget::item {
                color: #0f172a;
            }
            QGroupBox {
                color: #0f172a;
                font-weight: bold;
            }
        """)
        
        layout = QVBoxLayout(self)
        
        # Información de la venta
        info_group = QGroupBox("Información General")
        info_layout = QFormLayout()
        
        info_layout.addRow("Factura:", QLabel(self.sale_data['invoice_number']))
        info_layout.addRow("Fecha:", QLabel(self.sale_data['created_at'].strftime("%d/%m/%Y %H:%M")))
        info_layout.addRow("Cliente:", QLabel(self.sale_data['customer_name']))
        info_layout.addRow("Método de Pago:", QLabel(self.sale_data['payment_method']))
        info_layout.addRow("Estado:", QLabel(self.sale_data['status']))
        
        info_group.setLayout(info_layout)
        layout.addWidget(info_group)
        
        # Tabla de productos
        products_group = QGroupBox("Productos")
        products_layout = QVBoxLayout()
        
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Producto", "Precio Unit.", "Cantidad", "Subtotal"])
        table.setRowCount(len(self.items_data))
        
        for row, item in enumerate(self.items_data):
            table.setItem(row, 0, QTableWidgetItem(item['product_name']))
            table.setItem(row, 1, QTableWidgetItem(f"${item['unit_price']:.2f}"))
            table.setItem(row, 2, QTableWidgetItem(str(item['quantity'])))
            table.setItem(row, 3, QTableWidgetItem(f"${item['subtotal']:.2f}"))
        
        header = table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        
        products_layout.addWidget(table)
        products_group.setLayout(products_layout)
        layout.addWidget(products_group)
        
        # Totales
        totals_layout = QFormLayout()
        totals_layout.addRow("Subtotal:", QLabel(f"${self.sale_data['subtotal']:,.2f}"))
        totals_layout.addRow("Impuesto:", QLabel(f"${self.sale_data['tax']:,.2f}"))
        totals_layout.addRow("Descuento:", QLabel(f"${self.sale_data['discount']:,.2f}"))
        
        total_label = QLabel(f"${self.sale_data['total']:,.2f}")
        total_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #10b981;")
        totals_layout.addRow("TOTAL:", total_label)
        
        layout.addLayout(totals_layout)
        
        # Botón cerrar
        btn_close = QPushButton("Cerrar")
        btn_close.clicked.connect(self.accept)
        layout.addWidget(btn_close)


class EditSaleDialog(QDialog):
    def __init__(self, parent=None, sale=None):
        super().__init__(parent)
        self.sale = sale
        self.setWindowTitle(f"Editar Venta - {sale.invoice_number}")
        self.setMinimumWidth(400)
        layout = QVBoxLayout(self)
        form = QFormLayout()
        
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate(sale.created_at.year, sale.created_at.month, sale.created_at.day))
        form.addRow("Fecha:", self.date_edit)
        
        self.method_combo = QComboBox()
        for method in PaymentMethod:
            self.method_combo.addItem(method.value, method)
        # seleccionar
        idx = self.method_combo.findText(sale.payment_method.value)
        if idx >= 0:
            self.method_combo.setCurrentIndex(idx)
        form.addRow("Método de pago:", self.method_combo)
        
        self.tax_spin = QDoubleSpinBox()
        self.tax_spin.setPrefix("$ ")
        self.tax_spin.setMaximum(999999)
        self.tax_spin.setValue(float(sale.tax))
        form.addRow("Impuesto:", self.tax_spin)
        
        self.discount_spin = QDoubleSpinBox()
        self.discount_spin.setPrefix("$ ")
        self.discount_spin.setMaximum(999999)
        self.discount_spin.setValue(float(sale.discount))
        form.addRow("Descuento:", self.discount_spin)
        
        self.status_combo = QComboBox()
        for s in SaleStatus:
            self.status_combo.addItem(s.value, s)
        idxs = self.status_combo.findText(sale.status.value)
        if idxs >= 0:
            self.status_combo.setCurrentIndex(idxs)
        form.addRow("Estado:", self.status_combo)
        
        layout.addLayout(form)
        
        buttons = QHBoxLayout()
        btn_cancel = QPushButton("Cancelar")
        btn_cancel.clicked.connect(self.reject)
        btn_save = QPushButton("Guardar cambios")
        btn_save.clicked.connect(self.save)
        buttons.addStretch()
        buttons.addWidget(btn_cancel)
        buttons.addWidget(btn_save)
        layout.addLayout(buttons)

    def save(self):
        session = get_session()
        try:
            sale = session.query(Sale).filter_by(id=self.sale.id).first()
            old_status = sale.status
            new_status = self.status_combo.currentData()
            
            # Si se está cancelando la venta, devolver inventario
            if old_status != SaleStatus.CANCELLED and new_status == SaleStatus.CANCELLED:
                # Confirmar cancelación
                reply = QMessageBox.question(
                    self,
                    "Confirmar Cancelación",
                    "¿Está seguro de cancelar esta venta?\n\n"
                    "Se devolverán al inventario:\n"
                    "• Los productos vendidos\n"
                    "• Las materias primas utilizadas",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                
                if reply != QMessageBox.StandardButton.Yes:
                    return
                
                # Devolver productos al inventario
                for item in sale.items:
                    product = session.query(Product).filter_by(id=item.product_id).first()
                    if product:
                        previous_stock = product.stock
                        product.stock += item.quantity
                        
                        # Buscar y anular el movimiento de venta original
                        original_movement = session.query(InventoryMovement).filter(
                            InventoryMovement.reference == f"SALE-{sale.id}",
                            InventoryMovement.product_id == product.id
                        ).first()
                        
                        if original_movement:
                            # Anular el movimiento original agregando nota
                            original_movement.note = f"[ANULADO] Cancelación de venta {sale.invoice_number}"
                        
                        # Registrar movimiento de devolución
                        movement = InventoryMovement(
                            product_id=product.id,
                            movement_type=MovementType.ENTRY,
                            quantity=item.quantity,
                            previous_stock=previous_stock,
                            new_stock=product.stock,
                            reason=f"Devolución por cancelación de venta {sale.invoice_number}",
                            reference=f"CANCELLED-SALE-{sale.id}"  # Referencia a venta cancelada
                        )
                        session.add(movement)
                        
                        # Devolver materias primas
                        product_materials = session.query(ProductMaterial).filter_by(product_id=product.id).all()
                        for pm in product_materials:
                            raw_material = session.query(RawMaterial).filter_by(id=pm.raw_material_id).first()
                            if raw_material:
                                quantity_to_return = pm.quantity_needed * item.quantity
                                previous_material_stock = raw_material.stock
                                raw_material.stock += quantity_to_return
                                
                                # Registrar movimiento de materia prima
                                material_movement = RawMaterialMovement(
                                    raw_material_id=raw_material.id,
                                    movement_type=RawMaterialMovementType.RETURN,
                                    quantity=quantity_to_return,
                                    reference=sale.invoice_number,
                                    reason=f"Devolución por cancelación de venta {sale.invoice_number}"
                                )
                                session.add(material_movement)
            
            sale.payment_method = self.method_combo.currentData()
            sale.tax = self.tax_spin.value()
            sale.discount = self.discount_spin.value()
            sale.status = new_status
            # actualizar fecha (conservando hora original)
            new_date = self.date_edit.date().toPyDate()
            sale.created_at = sale.created_at.replace(year=new_date.year, month=new_date.month, day=new_date.day)
            sale.calculate_total()
            session.commit()
            
            if new_status == SaleStatus.CANCELLED:
                QMessageBox.information(self, "Venta cancelada", "La venta ha sido cancelada y el inventario ha sido devuelto")
            else:
                QMessageBox.information(self, "Venta modificada", "Los datos básicos de la venta fueron actualizados")
            self.accept()
        except Exception as e:
            session.rollback()
            QMessageBox.critical(self, "Error", f"No se pudo guardar: {str(e)}\n\n{traceback.format_exc()}")
        finally:
            close_session()


class EditSaleFullDialog(QDialog):
    """Diálogo para editar una venta completa (items, cantidades, etc.)"""
    def __init__(self, parent=None, sale_id=None):
        super().__init__(parent)
        self.sale_id = sale_id
        self.sale_items = []
        self.original_items = []  # Para saber qué devolver al inventario
        self.load_sale_data()
        self.init_ui()
        # Foco en el campo de pago
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(100, self.set_focus_to_payment)
    
    def load_sale_data(self):
        """Carga los datos de la venta a editar"""
        session = get_session()
        try:
            from sqlalchemy.orm import joinedload
            sale = session.query(Sale).options(joinedload(Sale.items)).filter_by(id=self.sale_id).first()
            if not sale:
                QMessageBox.warning(self, "Error", "Venta no encontrada")
                return
            
            self.sale_invoice = sale.invoice_number
            self.sale_customer_id = sale.customer_id
            self.sale_payment_method = sale.payment_method
            self.sale_tax = sale.tax or 0.0  # Cargar impuesto de la venta
            self.sale_transfer_type = sale.transfer_type if sale.transfer_type else None  # Cargar tipo de transferencia
            
            # Cargar items originales
            for item in sale.items:
                item_data = {
                    'product_id': item.product_id,
                    'product_name': item.product.name,
                    'unit_price': float(item.unit_price),
                    'quantity': item.quantity,
                    'subtotal': float(item.subtotal)
                }
                self.sale_items.append(item_data.copy())
                self.original_items.append(item_data.copy())
        finally:
            close_session()
    
    def init_ui(self):
        self.setWindowTitle(f"Editar Venta - {self.sale_invoice}")
        # Hacer la ventana de pantalla completa
        self.setWindowState(Qt.WindowState.WindowMaximized)
        # También establecer tamaño mínimo para casos donde no se maximice
        self.setMinimumSize(900, 600)
        self.setStyleSheet("""
            QDialog {
                background-color: white;
            }
            QLabel {
                color: #0f172a;
                background-color: transparent;
            }
            QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox {
                color: #0f172a;
                background-color: white;
            }
            QTableWidget {
                color: #0f172a;
            }
            QTableWidget::item {
                color: #0f172a;
            }
            QGroupBox {
                color: #0f172a;
                font-weight: bold;
            }
            QToolTip {
                background-color: #ffffff;
                color: #0f172a;
                border: 2px solid #3b82f6;
                border-radius: 8px;
                padding: 10px;
                font-size: 13px;
                font-weight: normal;
            }
        """)
        
        # Layout principal del diálogo
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Crear un scroll area para todo el contenido
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: white;
            }
            QScrollBar:vertical {
                background-color: #f1f5f9;
                width: 14px;
                border-radius: 7px;
                margin: 2px;
            }
            QScrollBar::handle:vertical {
                background-color: #cbd5e1;
                border-radius: 6px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #94a3b8;
            }
            QScrollBar:horizontal {
                background-color: #f1f5f9;
                height: 14px;
                border-radius: 7px;
                margin: 2px;
            }
            QScrollBar::handle:horizontal {
                background-color: #cbd5e1;
                border-radius: 6px;
                min-width: 30px;
            }
            QScrollBar::handle:horizontal:hover {
                background-color: #94a3b8;
            }
        """)
        
        # Widget contenedor para el scroll
        scroll_widget = QWidget()
        layout = QVBoxLayout(scroll_widget)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)
        
        # Mensaje de advertencia removido
        
        # Sección superior - Info de la venta
        top_section = QGroupBox("Información de la Venta")
        top_layout = QFormLayout()
        
        # Cliente
        self.customer_combo = QComboBox()
        self.customer_combo.setMinimumHeight(35)
        self.customer_combo.setStyleSheet("""
            QComboBox {
                background-color: white;
                color: #0f172a;
                border: 2px solid #e2e8f0;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 13px;
            }
            QComboBox:focus {
                border-color: #3b82f6;
                outline: none;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #0f172a;
                border: 1px solid #e2e8f0;
                selection-background-color: #3b82f6;
                selection-color: white;
                padding: 4px;
            }
            QComboBox::drop-down {
                border: none;
                background-color: transparent;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 7px solid #64748b;
                margin-right: 5px;
            }
        """)
        self.load_customers()
        # Seleccionar cliente actual
        if self.sale_customer_id:
            for i in range(self.customer_combo.count()):
                if self.customer_combo.itemData(i) == self.sale_customer_id:
                    self.customer_combo.setCurrentIndex(i)
                    break
        top_layout.addRow("Cliente:", self.customer_combo)
        
        # Método de pago
        self.payment_method_combo = QComboBox()
        self.payment_method_combo.setMinimumHeight(35)
        for method in PaymentMethod:
            self.payment_method_combo.addItem(method.value, method)
        # Seleccionar método actual
        for i in range(self.payment_method_combo.count()):
            if self.payment_method_combo.itemData(i) == self.sale_payment_method:
                self.payment_method_combo.setCurrentIndex(i)
                break
        self.payment_method_combo.setStyleSheet("""
            QComboBox {
                background-color: white;
                color: #0f172a;
                border: 2px solid #e2e8f0;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 13px;
            }
            QComboBox:focus {
                border-color: #3b82f6;
                outline: none;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #0f172a;
                border: 1px solid #e2e8f0;
                selection-background-color: #3b82f6;
                selection-color: white;
                padding: 4px;
            }
            QComboBox::drop-down {
                border: none;
                background-color: transparent;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 7px solid #64748b;
                margin-right: 5px;
            }
        """)
        self.payment_method_combo.currentIndexChanged.connect(self.on_payment_method_changed)
        top_layout.addRow("Método de Pago:", self.payment_method_combo)
        
        # Tipo de transferencia (oculto inicialmente)
        self.transfer_type_label = QLabel("Tipo de Transferencia:")
        self.transfer_type_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #0f172a;")
        self.transfer_type_label.setVisible(False)
        
        self.transfer_type_combo = QComboBox()
        self.transfer_type_combo.setMinimumHeight(35)
        self.transfer_type_combo.addItems(["Nequi", "Daviplata", "Bancolombia", "Otro"])
        self.transfer_type_combo.setVisible(False)
        self.transfer_type_combo.setStyleSheet("""
            QComboBox {
                background-color: white;
                color: #0f172a;
                border: 2px solid #e2e8f0;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 13px;
            }
            QComboBox:focus {
                border-color: #3b82f6;
                outline: none;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                color: #0f172a;
                border: 1px solid #e2e8f0;
                selection-background-color: #3b82f6;
                selection-color: white;
                padding: 4px;
            }
            QComboBox::drop-down {
                border: none;
                background-color: transparent;
                width: 30px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 7px solid #64748b;
                margin-right: 5px;
            }
        """)
        self.transfer_type_combo.currentIndexChanged.connect(self.on_transfer_type_changed)
        top_layout.addRow(self.transfer_type_label, self.transfer_type_combo)
        
        # Campo para "Otro" tipo de transferencia
        self.other_transfer_label = QLabel("Especifique:")
        self.other_transfer_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #0f172a;")
        self.other_transfer_label.setVisible(False)
        
        self.other_transfer_input = QLineEdit()
        self.other_transfer_input.setMinimumHeight(35)
        self.other_transfer_input.setPlaceholderText("Ej: Paypal, Bitcoin, etc.")
        self.other_transfer_input.setVisible(False)
        self.other_transfer_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: #0f172a;
                border: 2px solid #e2e8f0;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #3b82f6;
                outline: none;
            }
        """)
        top_layout.addRow(self.other_transfer_label, self.other_transfer_input)
        
        top_section.setLayout(top_layout)
        layout.addWidget(top_section)
        
        # Sección de productos
        products_section = QGroupBox("Productos")
        products_layout = QVBoxLayout()
        
        # Galería de productos (reducida para pantallas pequeñas)
        self.products_scroll = QScrollArea()
        self.products_scroll.setWidgetResizable(True)
        self.products_scroll.setMinimumHeight(120)
        self.products_scroll.setMaximumHeight(150)
        self.products_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)  # Scroll vertical cuando sea necesario
        self.products_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)  # Scroll horizontal cuando sea necesario
        self.products_scroll.setStyleSheet("""
            QScrollArea {
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                background-color: #f8fafc;
            }
            QScrollBar:vertical {
                background-color: #f1f5f9;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #cbd5e1;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #94a3b8;
            }
            QScrollBar:horizontal {
                background-color: #f1f5f9;
                height: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:horizontal {
                background-color: #cbd5e1;
                border-radius: 6px;
                min-width: 20px;
            }
            QScrollBar::handle:horizontal:hover {
                background-color: #94a3b8;
            }
        """)
        products_container = QWidget()
        self.products_grid = QGridLayout(products_container)
        self.products_grid.setContentsMargins(5, 5, 5, 5)
        self.products_grid.setHorizontalSpacing(5)
        self.products_grid.setVerticalSpacing(5)
        self.products_scroll.setWidget(products_container)
        products_layout.addWidget(self.products_scroll)
        
        self.build_products_gallery()
        
        # Tabla de items
        items_label = QLabel("Items de la venta:")
        items_label.setStyleSheet("font-weight: bold; font-size: 14px; margin-top: 10px;")
        products_layout.addWidget(items_label)
        
        self.items_table = QTableWidget()
        self.items_table.setColumnCount(5)
        self.items_table.setHorizontalHeaderLabels(["Producto", "Precio Unit.", "Cantidad", "Subtotal", "Acciones"])
        self.items_table.horizontalHeader().setStretchLastSection(False)
        self.items_table.setMinimumHeight(150)  # Altura mínima reducida para pantallas pequeñas
        self.items_table.setMaximumHeight(300)  # Altura máxima para controlar el espacio en pantallas pequeñas
        self.items_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #e2e8f0;
                background-color: white;
                font-size: 13px;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #f1f5f9;
                color: #0f172a;
            }
            QTableWidget::item:selected {
                background-color: #dbeafe;
                color: #1e40af;
            }
            QHeaderView::section {
                background-color: #f8fafc;
                padding: 8px;
                border: none;
                border-bottom: 2px solid #e2e8f0;
                font-weight: bold;
                color: #0f172a;
            }
            QScrollBar:vertical {
                background-color: #f1f5f9;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #cbd5e1;
                border-radius: 6px;
                min-height: 20px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #94a3b8;
            }
            QScrollBar:horizontal {
                background-color: #f1f5f9;
                height: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:horizontal {
                background-color: #cbd5e1;
                border-radius: 6px;
                min-width: 20px;
            }
            QScrollBar::handle:horizontal:hover {
                background-color: #94a3b8;
            }
        """)
        
        # Configurar anchos de columnas
        header = self.items_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.items_table.setColumnWidth(1, 100)
        self.items_table.setColumnWidth(2, 80)
        self.items_table.setColumnWidth(3, 100)
        self.items_table.setColumnWidth(4, 160)  # Acciones (más ancho para dos botones)
        
        # Configurar scroll y comportamiento - AsNeeded para que aparezca dinámicamente
        self.items_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.items_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.items_table.setAlternatingRowColors(True)
        self.items_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.items_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        
        products_layout.addWidget(self.items_table)
        
        products_section.setLayout(products_layout)
        layout.addWidget(products_section)
        
        # Actualizar tabla con items cargados
        self.update_items_table()
        
        # Totales
        totals_layout = QHBoxLayout()
        
        totals_group = QGroupBox("Totales y Pago")
        totals_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 2px solid #e2e8f0;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: #0f172a;
            }
        """)
        totals_form = QFormLayout()
        totals_form.setSpacing(6)
        totals_form.setContentsMargins(10, 5, 10, 5)
        
        # Subtotal
        subtotal_label_title = QLabel("Subtotal:")
        subtotal_label_title.setStyleSheet("font-size: 13px; font-weight: bold; color: #0f172a;")
        
        self.subtotal_label = QLabel("$0")
        self.subtotal_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #64748b; background-color: #f8fafc; padding: 5px; border-radius: 4px;")
        totals_form.addRow(subtotal_label_title, self.subtotal_label)
        
        # Impuesto
        tax_label = QLabel("Impuesto:")
        tax_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #0f172a;")
        
        self.tax_input = QDoubleSpinBox()
        self.tax_input.setMinimum(0)
        self.tax_input.setMaximum(999999.99)
        self.tax_input.setDecimals(2)
        self.tax_input.setPrefix("$ ")
        self.tax_input.setMinimumHeight(28)
        self.tax_input.setMinimumWidth(100)
        self.tax_input.setValue(0.0)
        self.tax_input.setStyleSheet("""
            QDoubleSpinBox {
                font-size: 13px;
                padding: 4px;
                background-color: white;
                border: 2px solid #e2e8f0;
                border-radius: 4px;
                color: #0f172a;
            }
            QDoubleSpinBox:focus {
                border-color: #3b82f6;
            }
        """)
        self.tax_input.valueChanged.connect(self.calculate_total)
        totals_form.addRow(tax_label, self.tax_input)
        
        # Total
        total_label_title = QLabel("TOTAL:")
        total_label_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #0f172a;")
        
        self.total_label = QLabel("$0")
        self.total_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #10b981; background-color: #f0fdf4; padding: 6px; border-radius: 4px;")
        totals_form.addRow(total_label_title, self.total_label)
        
        pago_label = QLabel("Pago con:")
        pago_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #0f172a;")
        
        self.cash_given_edit = QLineEdit()
        self.cash_given_edit.setPlaceholderText("Ingrese el monto recibido")
        self.cash_given_edit.setMinimumHeight(28)
        self.cash_given_edit.setMinimumWidth(100)
        self.cash_given_edit.setStyleSheet("""
            QLineEdit {
                font-size: 14px; 
                padding: 4px; 
                background-color: #fef3c7; 
                border: 2px solid #fbbf24; 
                border-radius: 4px;
            }
            QLineEdit:focus {
                border-color: #d97706;
                background-color: #fef9c3;
            }
        """)
        self.cash_given_edit.textChanged.connect(self.on_payment_changed)
        totals_form.addRow(pago_label, self.cash_given_edit)
        
        cambio_label = QLabel("Cambio:")
        cambio_label.setStyleSheet("font-size: 13px; font-weight: bold; color: #0f172a;")
        
        self.change_label = QLabel("")
        self.change_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #64748b; background-color: #eff6ff; padding: 6px; border-radius: 4px; min-height: 30px;")
        totals_form.addRow(cambio_label, self.change_label)
        
        totals_group.setLayout(totals_form)
        totals_layout.addWidget(totals_group)
        
        layout.addLayout(totals_layout)
        
        # Cargar impuesto si existe (después de crear el campo tax_input)
        if hasattr(self, 'sale_tax'):
            self.tax_input.setValue(self.sale_tax)
        
        # Cargar tipo de transferencia si existe (después de crear los campos)
        if hasattr(self, 'sale_transfer_type'):
            # Verificar si el método de pago es Transferencia
            if self.sale_payment_method == PaymentMethod.TRANSFER:
                self.transfer_type_label.setVisible(True)
                self.transfer_type_combo.setVisible(True)
                
                if self.sale_transfer_type:
                    # Buscar el tipo en el combo o agregarlo si es "Otro"
                    transfer_type = self.sale_transfer_type
                    index = self.transfer_type_combo.findText(transfer_type)
                    if index >= 0:
                        self.transfer_type_combo.setCurrentIndex(index)
                    else:
                        # Si no está en la lista, es un tipo "Otro"
                        self.transfer_type_combo.setCurrentText("Otro")
                        self.other_transfer_input.setText(transfer_type)
                        self.other_transfer_label.setVisible(True)
                        self.other_transfer_input.setVisible(True)
                else:
                    # Si no hay tipo guardado, mostrar solo el selector
                    self.transfer_type_combo.setCurrentIndex(0)
            else:
                # Si el método de pago no es Transferencia, asegurar que los campos estén ocultos
                self.transfer_type_label.setVisible(False)
                self.transfer_type_combo.setVisible(False)
                self.other_transfer_label.setVisible(False)
                self.other_transfer_input.setVisible(False)
        
        # Calcular total inicial
        self.calculate_total()
        
        # Botones
        buttons_layout = QHBoxLayout()
        buttons_layout.setContentsMargins(0, 10, 0, 0)
        
        # Botón para cancelar la venta (devolver inventario)
        btn_cancel_sale = QPushButton("❌ Cancelar Venta")
        btn_cancel_sale.setMinimumHeight(36)
        btn_cancel_sale.setMinimumWidth(130)
        btn_cancel_sale.setStyleSheet("""
            QPushButton {
                background-color: #ef4444;
                color: white;
                font-weight: bold;
                font-size: 13px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #dc2626;
            }
        """)
        btn_cancel_sale.clicked.connect(self.cancel_sale)
        
        buttons_layout.addWidget(btn_cancel_sale)
        buttons_layout.addStretch()
        
        btn_cancel = QPushButton("Cerrar")
        btn_cancel.setMinimumHeight(36)
        btn_cancel.setMinimumWidth(100)
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #6b7280;
                color: white;
                font-weight: bold;
                font-size: 13px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #4b5563;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        
        btn_save = QPushButton("Guardar Cambios")
        btn_save.setMinimumHeight(36)
        btn_save.setMinimumWidth(130)
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b;
                color: white;
                font-weight: bold;
                font-size: 14px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background-color: #d97706;
            }
        """)
        btn_save.clicked.connect(self.save_sale)
        
        buttons_layout.addWidget(btn_cancel)
        buttons_layout.addWidget(btn_save)
        
        layout.addLayout(buttons_layout)
        
        # Agregar el widget al scroll area
        scroll_area.setWidget(scroll_widget)
        
        # Agregar el scroll area al layout principal
        main_layout.addWidget(scroll_area)
    
    def set_focus_to_payment(self):
        """Establece el foco en el campo de pago"""
        try:
            self.cash_given_edit.setFocus()
            self.cash_given_edit.selectAll()
        except Exception:
            pass
    
    def on_payment_method_changed(self):
        """Maneja el cambio en el método de pago"""
        current_method = self.payment_method_combo.currentData()
        
        # Mostrar campos de transferencia solo si se selecciona Transferencia
        is_transfer = current_method == PaymentMethod.TRANSFER
        
        self.transfer_type_label.setVisible(is_transfer)
        self.transfer_type_combo.setVisible(is_transfer)
        
        # Si no es transferencia, ocultar también el campo "Otro"
        if not is_transfer:
            self.other_transfer_label.setVisible(False)
            self.other_transfer_input.setVisible(False)
        else:
            # Si es transferencia, verificar si seleccionaron "Otro"
            self.on_transfer_type_changed()
    
    def on_transfer_type_changed(self):
        """Maneja el cambio en el tipo de transferencia"""
        # Mostrar campo "Otro" solo si se selecciona "Otro"
        show_other = self.transfer_type_combo.currentText() == "Otro"
        self.other_transfer_label.setVisible(show_other)
        self.other_transfer_input.setVisible(show_other)
        
        # Limpiar el campo si se cambia a otra opción
        if not show_other:
            self.other_transfer_input.clear()
    
    def get_transfer_type(self):
        """Obtiene el tipo de transferencia según la selección"""
        current_method = self.payment_method_combo.currentData()
        
        if current_method != PaymentMethod.TRANSFER:
            return None
        
        transfer_type = self.transfer_type_combo.currentText()
        if transfer_type == "Otro":
            # Usar el valor del campo de texto si está especificado
            other_text = self.other_transfer_input.text().strip()
            return other_text if other_text else "Otro"
        
        return transfer_type
    
    def load_customers(self):
        """Carga los clientes disponibles"""
        session = get_session()
        try:
            self.customer_combo.addItem("Cliente General", None)
            customers = session.query(Customer).all()
            for customer in customers:
                self.customer_combo.addItem(customer.name, customer.id)
        finally:
            close_session()
    
    def load_products(self):
        """Carga los productos disponibles"""
        session = get_session()
        try:
            from sqlalchemy.orm import joinedload
            # Cargar productos con sus categorías usando joinedload
            self._products_cache = session.query(Product).options(
                joinedload(Product.category)
            ).filter(Product.stock > 0).all()
            # Acceder a las relaciones ahora para que se carguen antes de cerrar la sesión
            for product in self._products_cache:
                _ = product.category  # Forzar carga de la relación
                _ = product.name  # Acceder a atributos básicos
        finally:
            close_session()
    
    def build_products_gallery(self):
        """Construye la grilla de botones con imagen para agregar rápido productos."""
        while self.products_grid.count():
            item = self.products_grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()
        
        if not hasattr(self, '_products_cache') or self._products_cache is None:
            self.load_products()
        
        max_cols = 10
        row = 0
        col = 0
        for product in (self._products_cache or [])[:30]:
            btn = QPushButton()
            btn.setFixedSize(90, 80)
            btn.setStyleSheet("""
                QPushButton { 
                    background-color: white; 
                    border: 1px solid #e2e8f0; 
                    border-radius: 6px; 
                    color: #0f172a;
                    font-size: 10px;
                    font-weight: bold;
                    text-align: center;
                }
                QPushButton:hover { 
                    background-color: #f1f5f9; 
                    border-color: #3b82f6;
                }
            """)
            # Cargar imagen del producto
            pix = None
            if product.image_path:
                # Construir ruta absoluta desde la raíz del proyecto
                base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
                image_path = os.path.join(base_dir, product.image_path)
                if os.path.exists(image_path):
                    pix = QPixmap(image_path)
            
            # Si no hay imagen o no se cargó, intentar rutas alternativas (retrocompatibilidad)
            if not pix or pix.isNull():
                # Intentar con SKU o nombre del producto
                base_dir = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
                alt_paths = [
                    os.path.join(base_dir, "resources", "images", f"{product.sku}.png"),
                    os.path.join(base_dir, "resources", "images", f"{product.name}.png"),
                ]
                for alt_path in alt_paths:
                    if os.path.exists(alt_path):
                        pix = QPixmap(alt_path)
                        if not pix.isNull():
                            break
            
            # Configurar imagen en el botón si existe
            if pix and not pix.isNull():
                # Escalar imagen para que quepa bien en la tarjeta (60x60 para tarjeta de 90x80)
                scaled_pix = pix.scaled(
                    60, 60,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation
                )
                icon = QIcon(scaled_pix)
                btn.setIcon(icon)
                btn.setIconSize(scaled_pix.size())
                btn.setText("")  # No mostrar texto cuando hay imagen
                # Configurar estilo del botón con imagen
                btn.setStyleSheet("""
                    QPushButton { 
                        background-color: white; 
                        border: 1px solid #e2e8f0; 
                        border-radius: 6px; 
                    }
                    QPushButton:hover { 
                        background-color: #f1f5f9; 
                        border-color: #3b82f6;
                        border-width: 2px;
                    }
                """)
            else:
                # Si no hay imagen, mostrar nombre y precio
                if product.sale_price == int(product.sale_price):
                    price_text = f"${int(product.sale_price):,}"
                else:
                    price_text = f"${product.sale_price:.2f}"
                
                # Dividir nombre largo en múltiples líneas
                display_name = product.name
                if len(product.name) > 12:
                    words = product.name.split()
                    if len(words) > 1:
                        mid_point = len(words) // 2
                        line1 = ' '.join(words[:mid_point])
                        line2 = ' '.join(words[mid_point:])
                        display_name = f"{line1}\n{line2}"
                    else:
                        mid_point = len(product.name) // 2
                        display_name = f"{product.name[:mid_point]}\n{product.name[mid_point:]}"
                
                btn.setText(f"{display_name}\n{price_text}")
            
            # Crear tooltip con información completa del producto
            category_name = product.category.name if product.category else "Sin categoría"
            if product.sale_price == int(product.sale_price):
                price_text = f"${int(product.sale_price):,}"
            else:
                price_text = f"${product.sale_price:.2f}"
            
            tooltip_text = (
                f"📦 {product.name}\n"
                f"🏷️ SKU: {product.sku}\n"
                f"💰 Precio: {price_text}\n"
                f"📊 Stock: {product.stock}\n"
                f"📁 Categoría: {category_name}\n"
            )
            if product.description:
                tooltip_text += f"\n📝 {product.description[:100]}..."
            tooltip_text += f"\n\n👆 Click agrega 1 unidad."
            
            btn.setToolTip(tooltip_text)
            btn.clicked.connect(lambda checked=False, p=product: self.quick_add_product(p, 1))
            self.products_grid.addWidget(btn, row, col)
            col += 1
            if col >= max_cols:
                col = 0
                row += 1
    
    def quick_add_product(self, product, units=1):
        """Agrega producto rápido"""
        # Verificar si el producto ya está en la lista
        for item in self.sale_items:
            if item['product_id'] == product.id:
                item['quantity'] += units
                item['subtotal'] = item['quantity'] * item['unit_price']
                self.update_items_table()
                self.calculate_total()
                return
        
        item_data = {
            'product_id': product.id,
            'product_name': product.name,
            'unit_price': product.sale_price,
            'quantity': units,
            'subtotal': product.sale_price * units
        }
        self.sale_items.append(item_data)
        self.update_items_table()
        self.calculate_total()
    
    def update_items_table(self):
        """Actualiza la tabla de items"""
        self.items_table.setRowCount(len(self.sale_items))
        
        for row, item in enumerate(self.sale_items):
            # Producto
            product_item = QTableWidgetItem(item['product_name'])
            product_item.setFont(QFont("Arial", 12, QFont.Weight.Medium))
            self.items_table.setItem(row, 0, product_item)
            
            # Precio unitario
            unit_price = item['unit_price']
            if unit_price == int(unit_price):
                price_text = f"${int(unit_price):,}"
            else:
                price_text = f"${unit_price:.2f}"
            price_item = QTableWidgetItem(price_text)
            price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            price_item.setFont(QFont("Arial", 11))
            self.items_table.setItem(row, 1, price_item)
            
            # Cantidad (editable)
            qty_item = QTableWidgetItem(str(item['quantity']))
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            qty_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            qty_item.setFlags(qty_item.flags() | Qt.ItemFlag.ItemIsEditable)
            self.items_table.setItem(row, 2, qty_item)
            
            # Subtotal
            subtotal = item['subtotal']
            if subtotal == int(subtotal):
                subtotal_text = f"${int(subtotal):,}"
            else:
                subtotal_text = f"${subtotal:.2f}"
            subtotal_item = QTableWidgetItem(subtotal_text)
            subtotal_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            subtotal_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
            self.items_table.setItem(row, 3, subtotal_item)
            
            # Botones de acciones: sumar y restar
            actions_widget = QWidget()
            actions_layout = QHBoxLayout(actions_widget)
            actions_layout.setContentsMargins(4, 4, 4, 4)
            actions_layout.setSpacing(4)
            
            # Botón sumar cantidad
            btn_increase = QPushButton("➕ +1")
            btn_increase.setFixedHeight(28)
            btn_increase.setFixedWidth(70)
            btn_increase.setStyleSheet("""
                QPushButton {
                    background-color: #10b981;
                    color: white;
                    font-weight: bold;
                    font-size: 11px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #059669;
                }
            """)
            btn_increase.clicked.connect(lambda checked, r=row: self.increase_item_quantity(r))
            actions_layout.addWidget(btn_increase)
            
            # Botón reducir cantidad
            btn_reduce = QPushButton("➖ -1")
            btn_reduce.setFixedHeight(28)
            btn_reduce.setFixedWidth(70)
            btn_reduce.setStyleSheet("""
                QPushButton {
                    background-color: #f59e0b;
                    color: white;
                    font-weight: bold;
                    font-size: 11px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #d97706;
                }
            """)
            btn_reduce.clicked.connect(lambda checked, r=row: self.reduce_item_quantity(r))
            actions_layout.addWidget(btn_reduce)
            
            self.items_table.setCellWidget(row, 4, actions_widget)
            
            self.items_table.setRowHeight(row, 40)
        
        # Conectar señal para detectar cambios en cantidad
        self.items_table.itemChanged.connect(self.on_quantity_changed)
    
    def increase_item_quantity(self, row):
        """Aumenta la cantidad de un item en 1"""
        if row < len(self.sale_items):
            # Necesitamos verificar stock disponible desde la base de datos
            session = get_session()
            try:
                from models import Product
                product_id = self.sale_items[row]['product_id']
                product = session.query(Product).filter_by(id=product_id).first()
                if product:
                    if self.sale_items[row]['quantity'] >= product.stock:
                        QMessageBox.warning(self, "Stock insuficiente", 
                            f"Solo hay {product.stock} unidades disponibles")
                        return
                    
                    self.sale_items[row]['quantity'] += 1
                    self.sale_items[row]['subtotal'] = self.sale_items[row]['quantity'] * self.sale_items[row]['unit_price']
                    self.update_items_table()
                    self.calculate_total()
            finally:
                close_session()
    
    def reduce_item_quantity(self, row):
        """Reduce la cantidad de un item en 1, o lo elimina si llega a 0"""
        if row < len(self.sale_items):
            if self.sale_items[row]['quantity'] > 1:
                self.sale_items[row]['quantity'] -= 1
                self.sale_items[row]['subtotal'] = self.sale_items[row]['quantity'] * self.sale_items[row]['unit_price']
            else:
                self.sale_items.pop(row)
            self.update_items_table()
            self.calculate_total()
    
    def on_quantity_changed(self, item):
        """Maneja el cambio manual de cantidad en la tabla"""
        if item.column() == 2:  # Columna de cantidad
            row = item.row()
            if row < len(self.sale_items):
                try:
                    new_quantity = int(item.text())
                    if new_quantity > 0:
                        self.sale_items[row]['quantity'] = new_quantity
                        self.sale_items[row]['subtotal'] = new_quantity * self.sale_items[row]['unit_price']
                        self.calculate_total()
                        # Actualizar solo el subtotal en la tabla
                        subtotal = self.sale_items[row]['subtotal']
                        if subtotal == int(subtotal):
                            subtotal_text = f"${int(subtotal):,}"
                        else:
                            subtotal_text = f"${subtotal:.2f}"
                        
                        subtotal_item = QTableWidgetItem(subtotal_text)
                        subtotal_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        subtotal_item.setFont(QFont("Arial", 11, QFont.Weight.Bold))
                        self.items_table.setItem(row, 3, subtotal_item)
                    elif new_quantity == 0:
                        # Eliminar el item si cantidad es 0
                        self.sale_items.pop(row)
                        self.update_items_table()
                        self.calculate_total()
                    else:
                        # Restaurar valor anterior si es negativo
                        item.setText(str(self.sale_items[row]['quantity']))
                except ValueError:
                    # Restaurar valor anterior si no es un número válido
                    item.setText(str(self.sale_items[row]['quantity']))
    
    def calculate_total(self):
        """Calcula los totales incluyendo impuesto"""
        subtotal = sum(item['subtotal'] for item in self.sale_items)
        tax = self.tax_input.value() if hasattr(self, 'tax_input') else 0.0
        total = subtotal + tax
        
        # Actualizar subtotal
        if subtotal == int(subtotal):
            formatted_subtotal = f"${int(subtotal):,}"
        else:
            formatted_subtotal = f"${subtotal:,.2f}"
        self.subtotal_label.setText(formatted_subtotal)
        
        # Actualizar total
        if total == int(total):
            formatted_total = f"${int(total):,}"
        else:
            formatted_total = f"${total:,.2f}"
        
        self.total_label.setText(formatted_total)
        self.update_change()
    
    def on_payment_changed(self):
        """Maneja el cambio en el campo de pago"""
        self.update_change()
    
    def update_change(self):
        """Actualiza el cambio"""
        try:
            subtotal = sum(item['subtotal'] for item in self.sale_items)
            tax = self.tax_input.value() if hasattr(self, 'tax_input') else 0.0
            total = subtotal + tax
            
            payment_text = self.cash_given_edit.text().strip()
            if not payment_text:
                cash_given = 0
            else:
                clean_text = payment_text.replace('$', '').replace(',', '').strip()
                try:
                    cash_given = float(clean_text) if clean_text else 0
                except ValueError:
                    cash_given = 0
            
            if cash_given == 0:
                self.change_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #64748b; background-color: #eff6ff; padding: 6px; border-radius: 4px; min-height: 30px;")
                self.change_label.setText("")
            else:
                change = cash_given - total
                color = "#10b981" if change >= 0 else "#ef4444"
                bg_color = "#f0fdf4" if change >= 0 else "#fee2e2"
                
                if change == int(change):
                    formatted_change = f"${int(change):,}"
                else:
                    formatted_change = f"${change:,.2f}"
                
                self.change_label.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {color}; background-color: {bg_color}; padding: 6px; border-radius: 4px; min-height: 30px;")
                self.change_label.setText(formatted_change)
        except Exception:
            pass
    
    def cancel_sale(self):
        """Cancela la venta y devuelve el inventario"""
        session = get_session()
        try:
            sale = session.query(Sale).filter_by(id=self.sale_id).first()
            if not sale:
                QMessageBox.critical(self, "Error", "Venta no encontrada")
                return
            
            # Verificar si ya está cancelada
            if sale.status == SaleStatus.CANCELLED:
                QMessageBox.warning(self, "Advertencia", "Esta venta ya está cancelada")
                return
            
            # Confirmar cancelación
            reply = QMessageBox.question(
                self,
                "❌ Confirmar Cancelación de Venta",
                f"¿Está seguro de cancelar la venta {self.sale_invoice}?\n\n"
                "Se devolverán al inventario:\n"
                "• Todos los productos vendidos\n"
                "• Todas las materias primas utilizadas\n\n"
                "Esta acción cambiará el estado de la venta a 'Cancelada'.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # Devolver productos e items originales al inventario
            for orig_item in self.original_items:
                product = session.query(Product).filter_by(id=orig_item['product_id']).first()
                if product:
                    previous_stock = product.stock
                    product.stock += orig_item['quantity']
                    
                    # Buscar y anular el movimiento de venta original
                    original_movement = session.query(InventoryMovement).filter(
                        InventoryMovement.reference == f"SALE-{sale.id}",
                        InventoryMovement.product_id == product.id
                    ).first()
                    
                    if original_movement:
                        # Anular el movimiento original agregando nota
                        original_movement.note = f"[ANULADO] Cancelación de venta {self.sale_invoice}"
                    
                    # Registrar movimiento de devolución
                    movement = InventoryMovement(
                        product_id=product.id,
                        movement_type=MovementType.ENTRY,
                        quantity=orig_item['quantity'],
                        previous_stock=previous_stock,
                        new_stock=product.stock,
                        reason=f"Devolución por cancelación de venta {self.sale_invoice}",
                        reference=f"CANCELLED-SALE-{sale.id}"  # Referencia a venta cancelada
                    )
                    session.add(movement)
                    
                    # Devolver materias primas
                    product_materials = session.query(ProductMaterial).filter_by(product_id=product.id).all()
                    for pm in product_materials:
                        raw_material = session.query(RawMaterial).filter_by(id=pm.raw_material_id).first()
                        if raw_material:
                            quantity_to_return = pm.quantity_needed * orig_item['quantity']
                            raw_material.stock += quantity_to_return
                            
                            material_movement = RawMaterialMovement(
                                raw_material_id=raw_material.id,
                                movement_type=RawMaterialMovementType.RETURN,
                                quantity=quantity_to_return,
                                reference=self.sale_invoice,
                                reason=f"Devolución por cancelación de venta {self.sale_invoice}"
                            )
                            session.add(material_movement)
            
            # Cambiar estado a cancelada
            sale.status = SaleStatus.CANCELLED
            session.commit()
            
            QMessageBox.information(
                self,
                "✅ Venta Cancelada",
                f"La venta {self.sale_invoice} ha sido cancelada exitosamente.\n\n"
                f"• Estado: Cancelada\n"
                f"• Inventario devuelto correctamente\n"
                f"• Materias primas devueltas"
            )
            self.accept()
            
        except Exception as e:
            session.rollback()
            QMessageBox.critical(self, "Error", f"Error al cancelar venta: {str(e)}\n\n{traceback.format_exc()}")
        finally:
            close_session()
    
    def save_sale(self):
        """Guarda los cambios en la venta"""
        if not self.sale_items:
            QMessageBox.warning(self, "Error", "Debe agregar al menos un producto")
            return
        
        session = get_session()
        try:
            sale = session.query(Sale).filter_by(id=self.sale_id).first()
            if not sale:
                QMessageBox.critical(self, "Error", "Venta no encontrada")
                return
            
            # Confirmar edición
            reply = QMessageBox.question(
                self,
                "Confirmar Edición",
                "¿Está seguro de editar esta venta?\n\n"
                "Se ajustará el inventario:\n"
                "• Se devolverán los productos originales\n"
                "• Se descontarán los nuevos productos\n"
                "• Se ajustarán las materias primas",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            # PASO 1: Devolver al inventario los productos originales
            for orig_item in self.original_items:
                product = session.query(Product).filter_by(id=orig_item['product_id']).first()
                if product:
                    previous_stock = product.stock
                    product.stock += orig_item['quantity']
                    
                    # Registrar movimiento
                    movement = InventoryMovement(
                        product_id=product.id,
                        movement_type=MovementType.ENTRY,
                        quantity=orig_item['quantity'],
                        previous_stock=previous_stock,
                        new_stock=product.stock,
                        reason=f"Devolución por edición de venta {self.sale_invoice}"
                    )
                    session.add(movement)
                    
                    # Devolver materias primas
                    product_materials = session.query(ProductMaterial).filter_by(product_id=product.id).all()
                    for pm in product_materials:
                        raw_material = session.query(RawMaterial).filter_by(id=pm.raw_material_id).first()
                        if raw_material:
                            quantity_to_return = pm.quantity_needed * orig_item['quantity']
                            raw_material.stock += quantity_to_return
                            
                            material_movement = RawMaterialMovement(
                                raw_material_id=raw_material.id,
                                movement_type=RawMaterialMovementType.RETURN,
                                quantity=quantity_to_return,
                                reference=self.sale_invoice,
                                reason=f"Devolución por edición de venta {self.sale_invoice}"
                            )
                            session.add(material_movement)
            
            # PASO 2: Eliminar items antiguos
            for item in sale.items:
                session.delete(item)
            session.flush()
            
            # PASO 3: Agregar nuevos items y descontar inventario
            for item_data in self.sale_items:
                # Verificar stock disponible
                product = session.query(Product).filter_by(id=item_data['product_id']).first()
                if not product:
                    raise Exception(f"Producto {item_data['product_name']} no encontrado")
                
                if product.stock < item_data['quantity']:
                    raise Exception(f"Stock insuficiente para {item_data['product_name']}. Disponible: {product.stock}")
                
                # Crear nuevo item
                item = SaleItem()
                item.product_id = item_data['product_id']
                item.quantity = item_data['quantity']
                item.unit_price = item_data['unit_price']
                item.subtotal = item_data['subtotal']
                sale.items.append(item)
                
                # Descontar stock
                previous_stock = product.stock
                product.stock -= item_data['quantity']
                
                # Registrar movimiento
                movement = InventoryMovement(
                    product_id=product.id,
                    movement_type=MovementType.EXIT,
                    quantity=-item_data['quantity'],
                    previous_stock=previous_stock,
                    new_stock=product.stock,
                    reason=f"Venta editada {self.sale_invoice}"
                )
                session.add(movement)
                
                # Descontar materias primas
                product_materials = session.query(ProductMaterial).filter_by(product_id=product.id).all()
                for pm in product_materials:
                    raw_material = session.query(RawMaterial).filter_by(id=pm.raw_material_id).first()
                    if raw_material:
                        quantity_to_deduct = pm.quantity_needed * item_data['quantity']
                        raw_material.stock -= quantity_to_deduct
                        
                        material_movement = RawMaterialMovement(
                            raw_material_id=raw_material.id,
                            movement_type=RawMaterialMovementType.PRODUCTION,
                            quantity=-quantity_to_deduct,
                            reference=self.sale_invoice,
                            reason=f"Producción venta editada {self.sale_invoice}"
                        )
                        session.add(material_movement)
            
            # Actualizar datos de la venta
            sale.customer_id = self.customer_combo.currentData()
            sale.payment_method = self.payment_method_combo.currentData()
            sale.tax = self.tax_input.value() if hasattr(self, 'tax_input') else 0.0  # Guardar impuesto
            sale.transfer_type = self.get_transfer_type()  # Guardar tipo de transferencia
            sale.discount = 0
            sale.status = SaleStatus.EDITED  # Marcar como editada
            sale.calculate_total()
            
            session.commit()
            
            QMessageBox.information(
                self,
                "Éxito",
                f"Venta editada correctamente\n\n"
                f"Factura: {self.sale_invoice}\n"
                f"Nuevo Total: ${sale.total:,.2f}\n"
                f"Estado: Editada"
            )
            self.accept()
            
        except Exception as e:
            session.rollback()
            QMessageBox.critical(self, "Error", f"Error al editar venta: {str(e)}\n\n{traceback.format_exc()}")
        finally:
            close_session()